"""Stage 96 — import Amy's newer workbooks, and cross-check them against this pipeline.

Two workbooks arrived on 2026-07-27:

  * `Hillsborough_GF_Trend_Schedules_FY18_FY27_v5_Audit_Edition.xlsx` (27 sheets)
  * `Hillsborough_Workbook_B_Fiscal_Sustainability_Risk_Model.xlsx` (20 sheets)

They matter more than their file size suggests, because they contain the two things she
had asked *this* project to build:

  * **Material Change Drivers** — a commentary layer. One row per driver with the amount,
    the period, the budget category, why it matters, the source, a confidence rating and
    the follow-up question to put to the town. That is the "commentary on the drivers of
    material changes" she asked for, written by her.

  * **Tax Equivalent Exposure** — every major obligation converted to cents per $100 of
    valuation, so commitments of different kinds sit in one comparable unit. In the last
    handoff email this pipeline proposed building exactly that. She had already built it,
    using the same $240,000-per-penny figure taken from the town's own budget.

So this stage does not rebuild any of it. It **imports and verifies**, the same contract
as stage 85: her workbooks are read, never written, and every figure that has an
independently-extracted counterpart in this pipeline is compared against it.

That comparison is the real value. Two people reading the same public documents by
different methods — her by hand in Excel, this pipeline by parsing the PDFs — is a
genuine independent check. Where they agree, the figure is corroborated twice over. Where
they disagree, one of us is wrong and the disagreement is reported rather than averaged,
hidden, or silently resolved in this pipeline's favour.

Her own arithmetic is checked too: a cents-per-$100 figure must equal its dollar amount
divided by the stated penny assumption, and a total must equal its parts.
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import DATASETS, SOURCES, doc_id_for_filename, read_json, write_json  # noqa: E402

warnings.filterwarnings("ignore")

FOLDER = ("Orange County Efficiency & Accountability Initiative/"
          "06 Budget & Financial Analysis - Hillsborough")
TRENDS = "Hillsborough_GF_Trend_Schedules_FY18_FY27_v5_Audit_Edition.xlsx"
WB_B = "Hillsborough_Workbook_B_Fiscal_Sustainability_Risk_Model.xlsx"

# Where a figure in her workbooks has an independently-extracted counterpart here.
# The comparison is the point of this stage, so the mapping is explicit rather than
# guessed by name similarity.
CROSSCHECKS = [
    ("General Fund Expenditures", 2027, "general_fund_expenditures", 2027),
    ("Sales Tax Revenue", 2027, None, None),          # no counterpart yet; imported only
    ("Debt Service", 2027, None, None),
]
PENNY_METRIC = "revenue_per_cent_of_tax_rate"


def sheet_rows(ws, header_contains: str, max_scan: int = 8):
    """Return (header, rows) for a sheet whose table starts below a title block.

    Her sheets open with a title and a Purpose line, so the header row is not row 1 and
    its position differs per sheet. It is located by content instead of by index.
    """
    header, start = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(v.lower() == header_contains.lower() for v in vals):
            header, start = vals, i + 1
            break
    if header is None:
        return None, []
    ncol = max((i for i, v in enumerate(header, 1) if v), default=0)
    rows = []
    for row in ws.iter_rows(min_row=start, max_col=ncol, values_only=True):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        rows.append({header[i]: vals[i] for i in range(ncol) if header[i]})
    return header[:ncol], rows


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def main() -> None:
    tpath, bpath = SOURCES / FOLDER / TRENDS, SOURCES / FOLDER / WB_B
    for p in (tpath, bpath):
        if not p.exists():
            sys.exit(f"missing {p}")
    source_docs = [doc_id_for_filename(n) for n in (TRENDS, WB_B)]
    if not all(source_docs):
        sys.exit(f"workbook(s) not in documents.json — run s00 first: "
                 f"{[n for n, i in zip((TRENDS, WB_B), source_docs) if not i]}")

    facts = read_json(DATASETS / "facts.json")["facts"]

    # The same fiscal year is often reported by more than one document — this year's
    # budget states FY2027, and last year's budget PROJECTED FY2027. Taking whichever
    # row happened to come last accused Amy of a 78% vs 68.3% disagreement on fund
    # balance when in fact she had the current figure and this comparison had reached
    # for last year's projection. Prefer the firmest basis, then the newest document.
    BASIS_RANK = {"actual": 0, "audited": 0, "budget": 1, "recommended": 1,
                  "adopted": 1, "estimate": 2, "projected": 3}

    def readings(metric, fy=None):
        rows = [f for f in facts if f["metric"] == metric
                and (fy is None or f["fiscal_year"] == fy)]
        return sorted(rows, key=lambda f: (BASIS_RANK.get(f.get("basis"), 9),
                                           -(f.get("fiscal_year") or 0),
                                           str(f.get("source_doc"))))

    def fact(metric, fy=None):
        r = readings(metric, fy)
        return r[0] if r else None

    penny_f = fact(PENNY_METRIC)
    penny = penny_f["value"] if penny_f else None

    tw = openpyxl.load_workbook(tpath, data_only=True, read_only=True)
    bw = openpyxl.load_workbook(bpath, data_only=True, read_only=True)

    problems: list[str] = []

    # ---- her commentary layer -------------------------------------------------
    _, drivers = sheet_rows(tw["Material_Change_Drivers"], "Driver / Topic")
    material = []
    for r in drivers:
        amount = num(r.get("Amount / Impact"))
        material.append({
            "driver": r.get("Driver / Topic"), "amount": amount,
            "period": r.get("Period"), "budget_category": r.get("Budget Category"),
            "commentary": r.get("Commentary / Why it matters"),
            "source": r.get("Source / Footnote"),
            "confidence": r.get("Confidence"),
            "follow_up_question": r.get("Follow-up Question"),
            # Her own unit, recomputed here so the site never has to do the arithmetic.
            "cents_equivalent": round(amount / penny, 4) if amount and penny else None,
        })

    # ---- her comparable-unit view, with her arithmetic verified ---------------
    _, exposure = sheet_rows(bw["12 Tax Equivalent Exposure"], "Item")
    tax_equiv, arith = [], []
    for r in exposure:
        amount, cents = num(r.get("Annual Cost")), num(r.get("Cents per $100 Equivalent"))
        assumed = num(r.get("Tax Penny Assumption"))
        rec = {"item": r.get("Item"), "annual_cost": amount,
               "cents_per_100": round(cents, 4) if cents is not None else None,
               "penny_assumption": assumed, "status": r.get("Status"),
               "notes": r.get("Notes")}
        if amount and cents and assumed:
            expected = amount / assumed
            ok = abs(expected - cents) < 0.01
            arith.append({"item": r.get("Item"), "amount": amount,
                          "her_cents": round(cents, 4), "recomputed": round(expected, 4),
                          "agrees": ok})
            if not ok:
                problems.append(f"Tax Equivalent Exposure: '{r.get('Item')}' states "
                                f"{cents:.4f} cents but {amount:,.0f}/{assumed:,.0f} = "
                                f"{expected:.4f}")
        # Does her penny assumption match the town's published figure that this
        # pipeline extracted independently?
        if assumed and penny:
            rec["penny_matches_published"] = abs(assumed - penny) < 1.0
            if not rec["penny_matches_published"]:
                problems.append(f"her penny assumption {assumed:,.0f} differs from the "
                                f"town's published {penny:,.0f}")
        tax_equiv.append(rec)

    # ---- the FY29 cliff, and whether her total equals its parts ---------------
    _, cliff_rows = sheet_rows(bw["13 FY29 Fiscal Cliff"], "Component")
    cliff, cliff_total_stated, parts = [], None, 0.0
    for r in cliff_rows:
        amt = num(r.get("Annual Amount"))
        item = {"component": r.get("Component"), "annual_amount": amt,
                "cents": round(num(r.get("Tax Equivalent Cents")), 4)
                         if num(r.get("Tax Equivalent Cents")) is not None else None,
                "status": r.get("Status"), "notes": r.get("Notes")}
        if str(r.get("Component") or "").lower().startswith("total"):
            cliff_total_stated = amt
        else:
            cliff.append(item)
            if amt:
                parts += amt
    total_ok = (cliff_total_stated is not None
                and abs(parts - cliff_total_stated) < 1.0)
    if cliff_total_stated is not None and not total_ok:
        problems.append(f"FY29 Fiscal Cliff: parts sum to {parts:,.0f} but the total row "
                        f"states {cliff_total_stated:,.0f}")

    # ---- sales tax history, which this pipeline had no source for -------------
    _, st_rows = sheet_rows(tw["Sales_Tax_Detail"], "Fiscal Year")
    sales_tax = []
    for r in st_rows:
        fy = str(r.get("Fiscal Year") or "")
        amt = num(r.get("Sales Tax"))
        if not fy.startswith("FY") or amt is None:
            continue
        sales_tax.append({"fiscal_year": 2000 + int(fy[2:]), "amount": amt,
                          "source": r.get("Source"), "notes": r.get("Notes")})

    # ---- policy shifts: categories whose share changed materially ------------
    _, ps_rows = sheet_rows(bw["2 Policy Shifts"], "Category")
    policy = [{"category": r.get("Category"),
               "by_year": {y: num(r.get(y)) for y in ("FY18", "FY22", "FY25", "FY27")
                           if num(r.get(y)) is not None},
               "out_year": num(r.get("FY29 / Out-year")),
               "note": r.get("Interpretive Note")}
              for r in ps_rows if r.get("Category")]

    # ---- the cross-check that makes two independent readings worth having ----
    _, dash = sheet_rows(bw["0 Executive Dashboard"], "Key Metric")
    compare = []
    DASH_MAP = {
        "General Fund Expenditures / Budget": ("general_fund_expenditures", "FY27", 2027),
        "Fund Balance % of Expenditures": ("general_fund_balance_pct_of_expenditures",
                                           "FY27", 2027),
    }
    for r in dash:
        key = str(r.get("Key Metric") or "").strip()
        if key not in DASH_MAP:
            continue
        metric, col, fy = DASH_MAP[key]
        hers = num(r.get(col))
        mine_f = fact(metric, fy)
        if hers is None or not mine_f:
            continue
        mine = mine_f["value"]
        # Her fund-balance figure is a ratio; this pipeline stores it as a percentage.
        if "pct" in metric and hers < 5:
            hers *= 100
        diff = hers - mine
        all_readings = readings(metric, fy)
        compare.append({
            "her_label": key, "fiscal_year": fy, "metric": metric,
            "hers": round(hers, 2), "this_pipeline": round(mine, 2),
            "difference": round(diff, 2),
            "agrees": abs(diff) <= max(1.0, abs(mine) * 0.005),
            "my_source": f"{mine_f.get('source_doc')} p.{mine_f.get('source_page')}",
            "my_basis": mine_f.get("basis"),
            # Every reading is carried, so a comparison can never look decisive when the
            # documents themselves disagree about the same year.
            "all_readings_this_pipeline_holds": [
                {"value": round(r["value"], 2), "basis": r.get("basis"),
                 "source": f"{r.get('source_doc')} p.{r.get('source_page')}"}
                for r in all_readings],
            "documents_disagree_about_this_year": len(all_readings) > 1,
        })
    for c in compare:
        if not c["agrees"]:
            problems.append(f"her '{c['her_label']}' FY{c['fiscal_year']} = {c['hers']:,} but "
                            f"this pipeline read {c['this_pipeline']:,} from {c['my_source']}")

    tw.close()
    bw.close()

    write_json(DATASETS / "workbook_b.json", {
        "generated_by": "etl/s96_workbook_b.py",
        "imported_from": [TRENDS, WB_B],
        # Manifest ids for the two workbooks, so the documents these figures trace to
        # are machine-joinable rather than bare filenames.
        "source_docs": source_docs,
        "arrived": "2026-07-27",
        "contract": ("Her workbooks are READ, never written — she edits them in Excel. Every "
                     "figure with an independently-extracted counterpart here is compared "
                     "against it, and disagreements are reported rather than resolved in this "
                     "pipeline's favour."),
        "why_this_matters": ("These two workbooks contain the commentary layer and the "
                            "comparable-unit view that this project had been asked to build. "
                            "Two independent readings of the same public documents — hers by "
                            "hand, this pipeline's by parsing — is a real check on both."),
        "verification": {
            "her_penny_assumption": penny,
            "published_penny_source": (f"{penny_f.get('source_doc')} p.{penny_f.get('source_page')}"
                                       if penny_f else None),
            "tax_equivalent_arithmetic": arith,
            "tax_equivalent_all_consistent": all(a["agrees"] for a in arith) if arith else None,
            "fy29_cliff_parts_sum": round(parts, 2),
            "fy29_cliff_total_stated": cliff_total_stated,
            "fy29_cliff_total_reconciles": total_ok,
            "cross_checks_against_this_pipeline": compare,
            "cross_checks_all_agree": all(c["agrees"] for c in compare) if compare else None,
        },
        "material_change_drivers": material,
        "tax_equivalent_exposure": tax_equiv,
        "fy29_fiscal_cliff": cliff,
        "policy_shifts": policy,
        "sales_tax_history": sales_tax,
        "sales_tax_caveat": ("This is the TOWN's local option sales tax REVENUE. It is not the "
                            "county's sales tax rate for schools, which is a different thing and "
                            "still has no source in the archive."),
        "her_open_questions_to_the_town": [
            {"topic": r.get("Topic"), "question": r.get("Question")}
            for r in sheet_rows(openpyxl.load_workbook(tpath, data_only=True,
                                                       read_only=True)["Questions_for_Town"],
                                "Topic")[1] if r.get("Topic")],
        "problems": problems,
    })

    print(f"  imported {len(material)} material change drivers, {len(tax_equiv)} tax-equivalent "
          f"items, {len(policy)} policy shifts, {len(sales_tax)} years of sales tax")
    print(f"  her penny assumption ${penny:,.0f}" if penny else "  no penny figure")
    print(f"  her tax-equivalent arithmetic consistent: "
          f"{all(a['agrees'] for a in arith) if arith else 'n/a'} ({len(arith)} checked)")
    print(f"  FY29 cliff parts ${parts:,.0f} vs stated ${cliff_total_stated or 0:,.0f} "
          f"-> reconciles {total_ok}")
    print(f"\n  cross-checks against this pipeline's own extraction:")
    for c in compare:
        flag = "agrees" if c["agrees"] else "DISAGREES"
        print(f"      {flag:9} {c['her_label'][:40]:42} hers {c['hers']:>12,.2f}  "
              f"mine {c['this_pipeline']:>12,.2f}")
    print(f"\n  her largest drivers, in the shared unit:")
    for m in sorted([m for m in material if m["amount"]], key=lambda x: -x["amount"])[:6]:
        print(f"      ${m['amount']:>10,.0f}  {m['cents_equivalent']:>5.2f}c  "
              f"{(m['driver'] or '')[:44]:46} [{m['confidence']}]")
    if problems:
        print(f"\n  {len(problems)} problem(s):")
        for p in problems[:8]:
            print(f"      {p}")


if __name__ == "__main__":
    main()
