"""Stage 102 — catalogue every Excel workbook, so Amy does not have to open them one by one.

Her problem, verbatim:

    "I have so many excel files that we were WIP and never completed. I don't have a log to
     know what the purpose of each file is... sometimes a new file was started to represent
     the expanded definition of the project scope. Some files were only Hillsborough, others
     for Orange County. It will take me so much time to review each file and sort out what
     to keep and build on."

So this reads each workbook and writes the log that does not exist: what it contains, which
government it covers, which years, which family it belongs to, and whether a later file
supersedes it.

**Her own suggestion is the key to it** — "looking at the file attributes would give some
clues with the date the file was created". Filesystem dates on a Drive export are the export
date, not the authoring date, so they are useless here. The authoring dates come from Google
Drive's `modifiedTime`, retrieved 2026-07-28 and recorded in DRIVE below. Ordering the
workbooks by that turns a pile of files into a visible sequence of decisions.

**Nothing is deleted and nothing is judged as worthless.** A superseded workbook is evidence
of how the thinking developed, and her Design Manual explicitly wants "lessons learned"
retained. The recommendation column says which file to BUILD ON, not which to throw away.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import DATASETS, SOURCES, write_json  # noqa: E402

# Authoring dates + sizes from Amy's Drive folder "Excel workbooks"
# (1CJtv2xsXCi7q4Rr9VSFmTTclwyPOYNAB), read 2026-07-28. `createdTime` in Drive is when she
# copied each file into the new folder today; `modifiedTime` is when the work was actually
# done, which is the only one that tells a story.
DRIVE = {
    "Hillsborough_GF_Trend_Schedules_FY18_FY27.xlsx": ("2026-06-18T21:11:14Z", 27863),
    "Hillsborough_GF_Trend_Schedules_FY18_FY27_v2.xlsx": ("2026-06-18T21:11:11Z", 45498),
    "Hillsborough_GF_Trend_Schedules_FY18_FY27_v3.xlsx": ("2026-06-18T21:11:05Z", 66893),
    "Hillsborough_Workbook_B_Fiscal_Sustainability_Risk_Model.xlsx": ("2026-06-25T12:37:03Z", 524304),
    "Hillsborough_GF_Trend_Schedules_FY18_FY27_v4.xlsx": ("2026-06-25T12:37:25Z", 96637),
    "Hillsborough_GF_Trend_Schedules_FY18_FY27_v5_Audit_Edition.xlsx": ("2026-07-07T19:29:04Z", 82308),
    "Hillsborough_Municipal_Financial_Database_v1.xlsx": ("2026-07-07T20:17:56Z", 27300),
    "Municipal Finance Database - Hillsborough - v1.0.xlsx": ("2026-07-07T20:25:44Z", 60575),
    "Municipal Financial Analysis - Hillsborough - v1.0.xlsx": ("2026-07-07T21:10:28Z", 29267),
    "Hillsborough_Municipal_Financial_Database_v2_FY18_FY29.xlsx": ("2026-07-07T21:11:01Z", 74205),
    "Municipal Finance Project Design Manual v1.0.xlsx": ("2026-07-07T23:38:53Z", 19923),
    "Orange_County_Municipal_Finance_Database_v1.0.xlsx": ("2026-07-08T13:41:01Z", 57333),
    "Orange_County_Municipal_Financial_Data_Warehouse_v1.2.xlsx": ("2026-07-08T16:02:56Z", 95789),
    "Orange_County_Municipal_Finance_Database_v1.1.xlsx": ("2026-07-08T17:09:36Z", 67896),
    "Orange_County_Municipal_Financial_Data_Warehouse_v2.0.xlsx": ("2026-07-14T22:10:06Z", 112197),
    "Orange_County_Municipal_Financial_Information_System_v2.2_Foundation.xlsx":
        ("2026-07-26T21:40:57Z", 215919),
}

# Which lineage a file belongs to. A "family" is a chain of files that are versions of one
# idea; the newest in a family supersedes the rest.
FAMILIES = [
    ("GF Trend Schedules",
     r"^Hillsborough_GF_Trend_Schedules",
     "Historical General Fund trend schedules for Hillsborough, FY18-FY27."),
    ("Hillsborough database",
     r"^(Hillsborough_Municipal_Financial_Database|Municipal Finance Database - Hillsborough|"
     r"Municipal Financial Analysis - Hillsborough)",
     "The Hillsborough municipal finance database — the schema Amy named."),
    ("Orange County database",
     r"^Orange_County_Municipal_(Finance_Database|Financial_Data_Warehouse|"
     r"Financial_Information_System)",
     "The Orange County line, which evolved into the v2.2 Foundation."),
    ("Risk model",
     r"^Hillsborough_Workbook_B",
     "Workbook B — the fiscal sustainability risk model. Not a version of anything else."),
    ("Design manual",
     r"^Municipal Finance Project Design Manual",
     "The project's own design manual: objectives, conventions, version history, open "
     "issues and the full Finance Director data request."),
]

HB = re.compile(r"hillsborough|ORG_HB", re.I)
OC = re.compile(r"orange[ _]county|ORG_OC", re.I)
YEAR = re.compile(r"\bFY\s?(\d{2,4})\b", re.I)


def _seconds_between(a: str, b: str):
    from datetime import datetime
    try:
        fmt = "%Y-%m-%dT%H:%M:%SZ"
        return abs(int((datetime.strptime(b, fmt) - datetime.strptime(a, fmt)).total_seconds()))
    except Exception:
        return None


def family_of(name: str) -> tuple[str, str]:
    for label, pat, desc in FAMILIES:
        if re.search(pat, name):
            return label, desc
    return "Unclassified", ""


def inspect(path: Path) -> dict:
    """Open a workbook and describe what is actually in it."""
    out = {"sheets": [], "rows_total": 0, "years": set(), "entities": set(),
           "has_source_ids": False, "has_confidence": False, "error": None}
    # NOT read_only. In read-only mode openpyxl trusts the stored <dimension> element, and
    # several of these workbooks do not carry a correct one — three files with hundreds of
    # populated cells reported max_row=1, i.e. ZERO rows. Reporting "your workbook is empty"
    # to the person who built it is the worst kind of wrong: confidently false, and it would
    # discredit the whole catalogue. These files are all under 1 MB, so pay the parse cost.
    # One workbook (OC Data Warehouse v1.2) fails a full parse but opens read-only, so it is
    # NOT corrupt — openpyxl chokes on something in its worksheet XML that Excel handles.
    # Falling back keeps it in the catalogue, with its counts flagged as approximate because
    # read-only mode trusts the stored dimension. Telling Amy her file is broken when Excel
    # opens it perfectly well would be worse than an approximate row count.
    wb = None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        out["parse_note"] = (f"Full parse failed ({type(e).__name__}) — read with a reduced "
                             f"reader instead. Row counts below are approximate. The file is "
                             f"not damaged; Excel opens it.")
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e2:
            out["error"] = f"{type(e2).__name__}: {e2}"
            out["years"], out["entities"] = [], []
            return out
    try:
        for ws in wb.worksheets:
            # Count rows that actually hold something, rather than trusting the dimension.
            if out.get("parse_note"):        # reduced reader: dimension-based, approximate
                rows = max(0, (ws.max_row or 1) - 1)
            else:
                rows = sum(1 for row in ws.iter_rows(values_only=True)
                           if any(v is not None and str(v).strip() != "" for v in row))
                rows = max(0, rows - 1)          # discount the header
            out["sheets"].append({"name": ws.title, "rows": rows,
                                  "cols": ws.max_column or 0})
            out["rows_total"] += rows
            # Scan a bounded window: enough to characterise, cheap on a big sheet.
            for row in ws.iter_rows(max_row=40, max_col=16, values_only=True):
                for v in row:
                    if v is None:
                        continue
                    s = str(v)
                    if "Source_ID" in s:
                        out["has_source_ids"] = True
                    if s.strip() == "Confidence":
                        out["has_confidence"] = True
                    if HB.search(s):
                        out["entities"].add("Hillsborough")
                    if OC.search(s):
                        out["entities"].add("Orange County")
                    for m in YEAR.finditer(s):
                        y = int(m.group(1))
                        out["years"].add(2000 + y if y < 100 else y)
    finally:
        wb.close()
    out["years"] = sorted(out["years"])
    out["entities"] = sorted(out["entities"])
    return out


def main() -> None:
    found = {}
    # sorted(): first-wins dedup by basename must not depend on filesystem order.
    for p in sorted(SOURCES.rglob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        found.setdefault(p.name, p)

    records = []
    for name, (authored, drive_size) in DRIVE.items():
        p = found.get(name)
        rec = {"workbook": name, "authored": authored[:10], "authored_utc": authored,
               "drive_size_bytes": drive_size, "held_locally": bool(p)}
        rec["family"], rec["family_description"] = family_of(name)
        if p:
            rec.update(inspect(p))
            rec["sheet_count"] = len(rec["sheets"])
        else:
            rec.update({"sheets": [], "sheet_count": None, "rows_total": None,
                        "years": [], "entities": [], "has_source_ids": None,
                        "has_confidence": None,
                        "note": "In Amy's Drive folder but not in this archive — needs an "
                                "export before it can be catalogued in full."})
        records.append(rec)

    # Anything on disk that is NOT in her folder is still worth listing.
    for name, p in sorted(found.items()):
        if name in DRIVE:
            continue
        rec = {"workbook": name, "authored": None, "authored_utc": None,
               "drive_size_bytes": None, "held_locally": True,
               "note": "In the project archive but NOT in Amy's 'Excel workbooks' folder."}
        rec["family"], rec["family_description"] = family_of(name)
        rec.update(inspect(p))
        rec["sheet_count"] = len(rec["sheets"])
        records.append(rec)

    records.sort(key=lambda r: (r["authored_utc"] or "9999", r["workbook"]))

    # ---- supersession, decided by date within a family -----------------------
    newest: dict[str, dict] = {}
    for r in records:
        if r["family"] == "Unclassified" or not r["authored_utc"]:
            continue
        cur = newest.get(r["family"])
        if cur is None or r["authored_utc"] > cur["authored_utc"]:
            newest[r["family"]] = r
    for r in records:
        head = newest.get(r["family"])
        if not head or not r["authored_utc"]:
            r["status"] = "unknown — no authoring date"
            r["recommendation"] = "Catalogue once exported"
        elif r is head:
            r["status"] = "CURRENT in its family"
            r["recommendation"] = "Build on this one"
        else:
            # Files saved minutes apart are one working session, not a decision to replace.
            # Four of the Hillsborough files land inside 54 minutes and two are 33 SECONDS
            # apart — there the date ordering is an artefact of the save order, and calling
            # one "superseded" would invent a judgement the evidence does not support.
            gap = _seconds_between(r["authored_utc"], head["authored_utc"])
            if gap is not None and gap < 900:
                r["status"] = f"same working session as {head['workbook']} ({gap // 60} min apart)"
                r["recommendation"] = ("Ordering is NOT evidence here — Amy should say which "
                                       "of these she intended to keep")
                r["same_session_as_head"] = True
            else:
                r["status"] = f"superseded by {head['workbook']}"
                r["recommendation"] = "Keep as history — do not build on"
                r["same_session_as_head"] = False

    fam_summary = {}
    for label, _, desc in FAMILIES:
        members = [r for r in records if r["family"] == label]
        if not members:
            continue
        head = newest.get(label)
        fam_summary[label] = {
            "description": desc,
            "workbooks": len(members),
            "first_authored": members[0]["authored"],
            "latest_authored": head["authored"] if head else None,
            "build_on": head["workbook"] if head else None,
            "superseded": [m["workbook"] for m in members
                           if m is not head and not m.get("same_session_as_head")],
            "same_session_ambiguity": [m["workbook"] for m in members
                                       if m.get("same_session_as_head")],
        }

    missing = [r["workbook"] for r in records if not r["held_locally"]]

    write_json(DATASETS / "workbook_audit.json", {
        "generated_by": "etl/s102_workbook_audit.py",
        "requested_by": ("Amy — \"I have so many excel files that we were WIP and never "
                         "completed. I don't have a log to know what the purpose of each file "
                         "is... It will take me so much time to review each file and sort out "
                         "what to keep and build on.\""),
        "method": ("Each workbook is opened and described — sheets, row counts, which "
                   "government it covers, which fiscal years it mentions, and whether it "
                   "carries Source_ID/Confidence columns. Files are ordered by their Google "
                   "Drive modifiedTime, which is when the work was done; the filesystem date "
                   "on an export is just the export date and tells you nothing."),
        "how_supersession_is_decided": ("Within a family, the most recently authored file is "
                                        "CURRENT and the rest are history. Nothing is deleted: "
                                        "a superseded workbook is the record of how the "
                                        "thinking developed, which her own Design Manual asks "
                                        "to keep."),
        "summary": {
            "workbooks_catalogued": len(records),
            "in_amys_drive_folder": len(DRIVE),
            "held_in_this_archive": sum(1 for r in records if r["held_locally"]),
            "awaiting_export": missing,
            "families": len(fam_summary),
            "authoring_span": [records[0]["authored"], records[-1]["authored"]]
                              if records and records[0]["authored"] else None,
        },
        "families": fam_summary,
        "workbooks": records,
    })

    # ---- the readable version — JSON is for the site, this is for Amy ---------
    md = ["# Workbook catalogue", "",
          "*Generated by `etl/s102_workbook_audit.py` — do not edit by hand.*", "",
          "Every Excel workbook in the project, what is in it, and which one to build on.",
          "Ordered by when the work was actually done (Google Drive's modified date — the",
          "date on an exported file is just the export date and tells you nothing).", "",
          f"**{len(records)} workbooks · {len(fam_summary)} families**", ""]
    for label, s_ in fam_summary.items():
        md += [f"## {label}", "", s_["description"], "",
               f"**Build on: `{s_['build_on']}`**", ""]
        members = [r for r in records if r["family"] == label]
        md += ["| Authored | Workbook | Sheets | Rows | Covers | Status |",
               "|---|---|---|---|---|---|"]
        for r in members:
            cov = ", ".join(r.get("entities") or []) or "—"
            yrs = r.get("years") or []
            if yrs:
                cov += f" · FY{min(yrs) % 100:02d}–FY{max(yrs) % 100:02d}"
            flag = "**CURRENT**" if r["status"].startswith("CURRENT") else (
                "⚠ same session" if r.get("same_session_as_head") else "history")
            md.append(f"| {r['authored'] or '?'} | `{r['workbook']}` | "
                      f"{r.get('sheet_count') or '?'} | {r.get('rows_total') or '?'} | {cov} | {flag} |")
        md.append("")
        amb = s_.get("same_session_ambiguity") or []
        if amb:
            md += ["> ⚠️ **Ordering is not evidence for these.** They were saved within minutes "
                   "of the current file — one working session, not a decision to replace. "
                   "Which one you meant to keep is your call, not the file dates':", ""]
            md += [f"> - `{m}`" for m in amb] + [""]
    notes = [r for r in records if r.get("parse_note")]
    if notes:
        md += ["## Files that need a note", ""]
        for r in notes:
            md += [f"- **`{r['workbook']}`** — {r['parse_note']}", ""]
    orphans = [r for r in records if not r.get("authored")]
    if orphans:
        md += ["## In the project archive but NOT in your 'Excel workbooks' folder", "",
               "Worth deciding whether these belong there too:", ""]
        md += [f"- `{r['workbook']}` — {r.get('sheet_count') or '?'} sheets, "
               f"{r.get('rows_total') or '?'} rows" for r in orphans] + [""]
    md += ["## What this does not tell you", "",
           "- It reads structure and size, not meaning. Two workbooks with the same tabs may",
           "  have been built for different purposes, and only you know which.",
           "- 'Build on' is the most recently authored file in each family. It is not a",
           "  judgement about which is best.",
           "- Nothing here should be deleted. A superseded workbook is the record of how the",
           "  thinking developed, which your own Design Manual asks to keep.", ""]
    out_md = Path(__file__).resolve().parent.parent / "docs" / "WORKBOOK_CATALOGUE.md"
    out_md.write_text("\n".join(md), encoding="utf-8")
    print(f"  wrote docs/WORKBOOK_CATALOGUE.md")

    print(f"  {len(records)} workbooks catalogued "
          f"({sum(1 for r in records if r['held_locally'])} held locally, "
          f"{len(missing)} awaiting export)\n")
    for label, s in fam_summary.items():
        print(f"  {label} — {s['workbooks']} files, {s['first_authored']} → {s['latest_authored']}")
        print(f"      BUILD ON: {s['build_on']}")
        for m in s["superseded"]:
            print(f"      history:  {m}")
        for m in s.get("same_session_ambiguity", []):
            print(f"      ⚠ same session, ordering is not evidence: {m}")
        print()
    if missing:
        print(f"  needs an export before it can be read ({len(missing)}):")
        for m in missing:
            print(f"      {m}")


if __name__ == "__main__":
    main()
