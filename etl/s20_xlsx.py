"""Stage 20 — parse the two spreadsheets.

Two very different artifacts:

  Issues Log.xlsx
      the initiative's working list of topics being researched.

  Hillsborough Data Request June 2027.xlsx
      a records request sent TO the Town. It is a *blank template* — 29 sheets
      of structure with only 7 populated values. Treating it as data would be a
      mistake; treating it as a scoreboard of what the Town has not yet answered
      is the honest and more useful reading, so that is what this emits.

Only two sheets carry real numbers, and both are attributed to a named
commissioner rather than to an audited statement. They are published with
extraction="stated" so nobody mistakes them for figures from an ACFR.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import (DATASETS, SOURCES, Fact, STATED, DERIVED,  # noqa: E402
                   write_json, parse_money)

REQUEST_XLSX = ("Orange County Efficiency & Accountability Initiative/"
                "08 Public Records Requests/Hillsborough Data Request June 2027.xlsx")
ISSUES_XLSX = ("Orange County Efficiency & Accountability Initiative/"
               "01 Master Issues List/Issues Log.xlsx")

REQUEST_DOC = "hillsborough-data-request-june-2027"
ISSUES_DOC = "issues-log"
JUR = "Town of Hillsborough, NC"

FY_RE = re.compile(r"^FY\s?(\d{2}|20\d{2})", re.I)

# Sheets that are section dividers, not data tables.
SECTION_SHEETS = {"1. Staffing & Compensation", "2. Utilities", "3. Capital Projects",
                  "4. Debt", "5. Revenue & Taxes", "6. Affordable Housing", "Request"}


def suffix_money(s: str) -> tuple[float | None, str]:
    """Parse '$791k', '$2.916M', '$1,124M' -> dollars.

    '$1,124M' is a typo in the source: read literally it is $1.124 billion for a
    greenway. The row's own arithmetic settles it ($791k + $333k = $1.124M), so
    we read it as $1,124k and say so in the note rather than silently choosing.
    """
    if s is None:
        return None, ""
    t = str(s).strip()
    m = re.match(r"^\$?\s*([\d,]+(?:\.\d+)?)\s*([kKmM])?\s*$", t)
    if not m:
        return None, ""
    raw, suf = m.group(1), (m.group(2) or "").lower()
    note = ""
    has_comma = "," in raw
    num = float(raw.replace(",", ""))
    if suf == "m" and has_comma:
        # "1,124M" — comma-grouped thousands wearing an M suffix.
        val = num * 1_000
        note = (f"Source cell reads {t!r}. Interpreted as ${num:,.0f} thousand "
                f"(= ${val:,.0f}) because the row's own Original+Increase "
                f"arithmetic requires it; the 'M' suffix in the source is a typo.")
    elif suf == "m":
        val = num * 1_000_000
    elif suf == "k":
        val = num * 1_000
    else:
        val = num
    return val, note


def parse_issues(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    header_i = next((i for i, r in enumerate(rows)
                     if r and any(isinstance(c, str) and c.strip() == "Issue" for c in r)), None)
    if header_i is None:
        return []
    header = [str(c).strip() if c else "" for c in rows[header_i]]
    col = {name: idx for idx, name in enumerate(header) if name}
    out = []
    for r in rows[header_i + 1:]:
        issue = r[col["Issue"]] if "Issue" in col and col["Issue"] < len(r) else None
        if not issue or not str(issue).strip():
            continue

        def g(name):
            i = col.get(name)
            if i is None or i >= len(r) or r[i] is None:
                return None
            v = str(r[i]).strip()
            return None if v in {"", "TBD"} else v

        out.append({
            "issue": str(issue).strip(),
            "status": g("Status"),
            "key_facts": g("Key Facts"),
            "source_documents": g("Source Documents"),
            "next_action": g("Next Action"),
            "source_doc": ISSUES_DOC,
            "source_detail": ws.title,
        })
    return out


def parse_request(path: Path) -> tuple[dict, list[Fact]]:
    """Read the request workbook as a fill-state scoreboard plus a few facts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    facts: list[Fact] = []

    cover = [str(c.value).strip() for r in wb["Request"].iter_rows()
             for c in r if c.value] if "Request" in wb.sheetnames else []

    section = None
    tables = []
    for ws in wb.worksheets:
        title = ws.title
        if title in SECTION_SHEETS:
            if title != "Request":
                section = re.sub(r"^\d+\.\s*", "", title)
            continue

        grid = [[c.value for c in r] for r in ws.iter_rows()]
        # Header row = the row with the most FY-looking cells.
        best_i, best_n = None, 0
        for i, r in enumerate(grid):
            n = sum(1 for c in r if isinstance(c, str) and FY_RE.match(c.strip()))
            if n > best_n:
                best_i, best_n = i, n
        # Otherwise the row that looks like column labels.
        if best_i is None:
            best_i = next((i for i, r in enumerate(grid)
                           if sum(1 for c in r if isinstance(c, str) and c.strip()) >= 2), 0)

        header = [(str(c).strip() if c is not None else "") for c in grid[best_i]]
        data_cols = [j for j, h in enumerate(header) if h and j > 0]
        row_labels, filled, expected = [], 0, 0
        for r in grid[best_i + 1:]:
            label = r[0] if r else None
            if not label or not str(label).strip():
                continue
            lab = str(label).strip()
            if lab.lower() in {"total", "department"} and not any(
                    r[j] is not None for j in data_cols if j < len(r)):
                continue
            row_labels.append(lab)
            for j in data_cols:
                expected += 1
                if j < len(r) and r[j] is not None and str(r[j]).strip() != "":
                    filled += 1

        status = "unanswered" if filled == 0 else ("answered" if filled >= expected else "partial")
        tables.append({
            "sheet": title,
            "section": section,
            "title": str(grid[0][0]).strip() if grid and grid[0] and grid[0][0] else title,
            "columns_requested": [h for h in header[1:] if h],
            "rows_requested": row_labels,
            "cells_expected": expected,
            "cells_provided": filled,
            "status": status,
        })

    # ---- the two sheets that do carry values -----------------------------
    ws = wb["Overview"]
    attribution = next((str(c.value) for r in ws.iter_rows() for c in r
                        if isinstance(c.value, str) and c.value.startswith("Source:")), "")
    series = []
    for row in ws.iter_rows():
        label = row[0].value
        if not isinstance(label, str) or not FY_RE.match(label.strip()):
            continue
        fy = 2000 + int(label.strip()[2:]) if len(label.strip()) == 4 else int(label.strip()[-4:])
        millions = row[1].value if len(row) > 1 else None
        if millions is None:
            continue
        facts.append(Fact(
            jurisdiction=JUR, fiscal_year=fy, metric="admin_spend_total",
            value=round(float(millions) * 1_000_000, 2), unit="USD", basis="stated",
            source_doc=REQUEST_DOC, source_detail=f"sheet 'Overview' row {label}",
            extraction=STATED,
            note=(f"Figure supplied in the initiative's request workbook and attributed "
                  f"there to {attribution.replace('Source: ', '')}. Not traced to an "
                  f"audited financial statement — treat as a claim to verify, not as "
                  f"an ACFR figure."),
        ))
        series.append((fy, float(millions)))

    # Year-over-year change we compute ourselves so the site never has to trust
    # a percentage typed into a spreadsheet.
    series.sort()
    for (y0, v0), (y1, v1) in zip(series, series[1:]):
        if v0:
            facts.append(Fact(
                jurisdiction=JUR, fiscal_year=y1, metric="admin_spend_yoy_pct",
                value=round((v1 - v0) / v0 * 100, 2), unit="percent", basis="derived",
                source_doc=REQUEST_DOC, source_detail="computed from admin_spend_total",
                extraction=DERIVED,
                note=f"Computed by this pipeline from FY{y0} and FY{y1} values.",
            ))
    if len(series) >= 2:
        (y0, v0), (y1, v1) = series[0], series[-1]
        facts.append(Fact(
            jurisdiction=JUR, fiscal_year=y1, metric="admin_spend_change_pct_since",
            value=round((v1 - v0) / v0 * 100, 2), unit="percent", basis="derived",
            source_doc=REQUEST_DOC, source_detail=f"FY{y0}->FY{y1}", extraction=DERIVED,
            note=f"Total change in stated administrative spend from FY{y0} to FY{y1}.",
        ))

    projects = []
    ws = wb["Project Cost Changes"]
    for row in ws.iter_rows():
        name = row[0].value
        if not isinstance(name, str) or name.strip() in {"", "Project", "Project Cost Changes"}:
            continue
        orig, n1 = suffix_money(row[1].value if len(row) > 1 else None)
        curr, n2 = suffix_money(row[2].value if len(row) > 2 else None)
        incr, n3 = suffix_money(row[3].value if len(row) > 3 else None)
        notes = " ".join(n for n in (n1, n2, n3) if n)
        checks = None
        if orig is not None and curr is not None and incr is not None:
            checks = abs((orig + incr) - curr) < 1_000  # self-consistent within $1k
        projects.append({
            "project": name.strip(),
            "original_budget_usd": orig,
            "current_budget_usd": curr,
            "increase_usd": incr,
            "increase_pct": round((curr - orig) / orig * 100, 1) if orig and curr else None,
            "arithmetic_consistent": checks,
            "raw_cells": [row[i].value if len(row) > i else None for i in (1, 2, 3)],
            "note": notes or None,
            "source_doc": REQUEST_DOC,
            "source_detail": "sheet 'Project Cost Changes'",
            "extraction": STATED,
        })
        if orig is not None:
            facts.append(Fact(jurisdiction=JUR, fiscal_year=None,
                              metric="capital_project_original_budget", value=orig, unit="USD",
                              basis="stated", source_doc=REQUEST_DOC,
                              source_detail=f"Project Cost Changes / {name.strip()}",
                              extraction=STATED, note=notes))
        if curr is not None:
            facts.append(Fact(jurisdiction=JUR, fiscal_year=None,
                              metric="capital_project_current_budget", value=curr, unit="USD",
                              basis="stated", source_doc=REQUEST_DOC,
                              source_detail=f"Project Cost Changes / {name.strip()}",
                              extraction=STATED, note=notes))

    unanswered = [t for t in tables if t["status"] == "unanswered"]
    scoreboard = {
        "request_document": REQUEST_DOC,
        "cover_note": cover,
        "summary": {
            "tables_requested": len(tables),
            "tables_unanswered": len(unanswered),
            "tables_partial": sum(1 for t in tables if t["status"] == "partial"),
            "tables_answered": sum(1 for t in tables if t["status"] == "answered"),
            "data_cells_requested": sum(t["cells_expected"] for t in tables),
            "data_cells_provided": sum(t["cells_provided"] for t in tables),
        },
        "tables": tables,
        "projects_with_cost_changes": projects,
    }
    return scoreboard, facts


def main() -> None:
    issues = parse_issues(SOURCES / ISSUES_XLSX)
    write_json(DATASETS / "issues.json", {
        "generated_by": "etl/s20_xlsx.py",
        "note": ("The initiative's working topic list. 'key_facts' of TBD means "
                 "research is open, not that no facts exist."),
        "issues": issues,
    })

    scoreboard, facts = parse_request(SOURCES / REQUEST_XLSX)
    write_json(DATASETS / "requests.json", {
        "generated_by": "etl/s20_xlsx.py",
        "note": ("A public records request sent to the Town of Hillsborough. The "
                 "workbook was sent as a blank template; the fill state below "
                 "reflects the copy in this archive and is a snapshot, not a "
                 "claim that the Town refused to respond."),
        **scoreboard,
    })

    write_json(DATASETS / "facts_xlsx.json", {
        "generated_by": "etl/s20_xlsx.py",
        "facts": [f.as_row() for f in facts],
    })

    s = scoreboard["summary"]
    print(f"\n  issues: {len(issues)}")
    print(f"  request tables: {s['tables_requested']} "
          f"({s['tables_unanswered']} unanswered, {s['tables_partial']} partial, "
          f"{s['tables_answered']} answered)")
    print(f"  data cells: {s['data_cells_provided']}/{s['data_cells_requested']} provided")
    print(f"  facts: {len(facts)}")


if __name__ == "__main__":
    main()
