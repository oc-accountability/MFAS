"""Stage 103 — the tab map: one workbook per row, every tab in its own column.

Amy's request:

    "Would you create an excel file that just documents what the 16 files contain? One file
     per row, and columns that contain a description of the file (or purpose), and then each
     tab in a different column. From there I hope to have a structure that I can more easily
     use to follow the different files."

Deliberately literal. She asked for a wide sheet she can read left-to-right, and a wide sheet
is what this makes — not a tidy long table, which would be better for a machine and worse for
the thing she actually wants to do, which is scan sixteen files and see the shape of them.

Two extra sheets earn their place beside it:

  * **Tab_Index** — the same information turned on its side: one row per TAB, so she can sort
    by tab name and immediately see which tabs recur across workbooks and which are unique.
    That is the view that answers "did I already solve this somewhere?", which is the question
    behind "I don't want to overlook something really smart that I had decided".

  * **Decisions_Inventory** — the conventions and ideas already embedded in those workbooks,
    harvested so they can be carried forward on purpose rather than by accident.

Row counts come from stage 102, which walks the rows rather than trusting the dimension
element — three of these workbooks under-report themselves as empty otherwise.
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import DATA, DATASETS, read_json  # noqa: E402

OUT = DATA / "exports" / "MFAS_Workbook_Tab_Map.xlsx"
HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(bold=True, color="FFFFFF", size=10)
TABF = PatternFill("solid", fgColor="E8EDF5")

# One line on what each workbook is FOR. Written by hand after reading them — a machine can
# list tabs but cannot say why a file exists, and "purpose" is the column she asked for.
PURPOSE = {
    "Hillsborough_GF_Trend_Schedules_FY18_FY27.xlsx":
        "First cut of the General Fund historical trend schedules. The starting point.",
    "Hillsborough_GF_Trend_Schedules_FY18_FY27_v2.xlsx":
        "Trend schedules, second pass — roughly doubles the content of v1.",
    "Hillsborough_GF_Trend_Schedules_FY18_FY27_v3.xlsx":
        "Trend schedules, third pass. Expense and revenue detail deepen.",
    "Hillsborough_GF_Trend_Schedules_FY18_FY27_v4.xlsx":
        "Big expansion — triples the row count and adds the analytical sheets.",
    "Hillsborough_GF_Trend_Schedules_FY18_FY27_v5_Audit_Edition.xlsx":
        "CURRENT trend schedules. Adds Material_Change_Drivers, Sales_Tax_Detail, "
        "reconciliation sheets and Questions_for_Town. The most complete Hillsborough history.",
    "Hillsborough_Workbook_B_Fiscal_Sustainability_Risk_Model.xlsx":
        "The risk model — executive dashboard, policy shifts, per-project bridges, tax "
        "equivalent exposure, the FY29 cliff, and base/optimistic/conservative cases.",
    "Hillsborough_Municipal_Financial_Database_v1.xlsx":
        "First Hillsborough star-schema database: Source_Register, Permanent_IDs, an "
        "Accounting_Roadmap and five fact tables. The schema the later work inherits.",
    "Municipal Finance Database - Hillsborough - v1.0.xlsx":
        "Same idea, restructured into eleven narrower fact tables (revenue, expense by "
        "function, by category, by type, staffing, debt, housing, capital, utility, stormwater).",
    "Municipal Financial Analysis - Hillsborough - v1.0.xlsx":
        "The ANALYSIS layer rather than the data layer — R1-R8 reserve, bridge, capacity and "
        "risk views, plus a glossary. A different kind of workbook from its neighbours.",
    "Hillsborough_Municipal_Financial_Database_v2_FY18_FY29.xlsx":
        "The fullest Hillsborough database — same eighteen-tab structure, populated, "
        "FY18-FY29 with forward years.",
    "Municipal Finance Project Design Manual v1.0.xlsx":
        "The project's own manual: objectives, naming and data standards, version history, "
        "lessons learned, open issues, and the complete eleven-section Finance Director data "
        "request. Not data — doctrine.",
    "Orange_County_Municipal_Finance_Database_v1.0.xlsx":
        "The Orange County line opens: county star schema, FY2025 fact tables.",
    "Orange_County_Municipal_Finance_Database_v1.1.xlsx":
        "County database with the fact tables filled out and a financial statement map.",
    "Orange_County_Municipal_Financial_Data_Warehouse_v1.2.xlsx":
        "Renamed to 'Data Warehouse' — the first attempt at historical series rather than a "
        "single year.",
    "Orange_County_Municipal_Financial_Data_Warehouse_v2.0.xlsx":
        "Clean production rebuild of the county warehouse, FY2018-FY2025 actuals.",
    "Orange_County_Municipal_Financial_Information_System_v2.2_Foundation.xlsx":
        "CURRENT county workbook. Adds a reading order, source catalog, review dashboard and "
        "the FY2025 walkthrough. Renamed again to 'Information System'.",
    "Issues Log.xlsx": "The master issues list for the initiative.",
    "Debt Service Projections.xlsx":
        "Debt schedules supplied by the Town's Finance Director in response to the records "
        "request — existing and projected debt service through maturity.",
    "Hillsborough Data Request June 2027.xlsx":
        "The records request TO the town, laid out as a fill-in workbook — 30 tabs of the "
        "schedules being asked for.",
    "OC Design.xlsx": "Early design notes for the Orange County side.",
}

# Conventions and ideas already present in her own workbooks. The point of listing them is
# her own: "I don't want to overlook something really smart that I had decided."
DECISIONS = [
    ("Permanent IDs with type prefixes",
     "ORG_ / FUND_ / REV_ / EXP_ / TYPE_ / STRAT_ / PROJ_",
     "Hillsborough_Municipal_Financial_Database_v1",
     "Keep. Type-prefixed keys are readable and sort into groups. STRAT_ (strategic "
     "initiative) and PROJ_ (capital project) are the two most valuable and appear nowhere "
     "else in the project."),
    ("Reserved IDs for governments not yet loaded",
     "ORG_CH, ORG_CB, ORG_MB declared with Status=Future",
     "Hillsborough_Municipal_Financial_Database_v1",
     "Keep — this is exactly the Chapel Hill question, already answered. The ID exists "
     "before the data does, so adding a municipality is loading rows, not redesigning."),
    ("Accounting_Roadmap",
     "Which accounting basis answers which question, and which figures may not be compared",
     "Hillsborough_Municipal_Financial_Database_v1",
     "Keep and promote. The clearest short statement of the hardest trap in the project: "
     "government-wide, governmental funds, budgetary and enterprise views are different "
     "things and mixing them silently produces wrong answers."),
    ("Source_ID + Source_Detail on every fact",
     "Machine key plus a human page reference",
     "Hillsborough database family",
     "Keep. Source_ID alone lets a figure be traced to a document; Source_Detail is what "
     "lets a person find it on the page."),
    ("Confidence vocabulary",
     "High / Medium / Working / Pending",
     "Orange County v2.2 Foundation",
     "Keep. Four levels is enough to be used consistently; more would not be."),
    ("Colour convention",
     "Yellow = pending town data or incomplete extract; red comment = high-priority gap",
     "Hillsborough_Municipal_Financial_Database_v1",
     "Keep for authored sheets, but do NOT rely on it for machine-read data — colour is "
     "invisible to every tool that reads the file. Pair it with a status column."),
    ("Scenario as a first-class field",
     "Actual / Adopted Budget / Recommended Budget / Estimate / Projection",
     "Hillsborough database family",
     "Keep, and this is the key to the FY2026-actuals question: a final actual is a NEW ROW "
     "with a different Scenario, not an overwrite of the estimate."),
    ("Data_Quality_Gaps as a standing sheet",
     "Gap_ID, Priority, Topic, Current Status, Why It Matters, Likely Resolution",
     "Both families",
     "Keep. A gap that is written down is a task; a gap that is remembered is a risk."),
    ("Lessons Learned",
     "e.g. 'Municipal budgets are legally balanced, not necessarily structurally balanced'",
     "Municipal Finance Project Design Manual",
     "Keep and expand. Her own note says these 'shouldn't disappear into old chats' — that "
     "is the single most valuable sentence in the sixteen files."),
    ("The eleven-section Finance Director request",
     "Fund balance, cost allocation, staffing, revenue, capital, debt, housing, assumptions, "
     "crosswalks, transfers, supporting documents",
     "Municipal Finance Project Design Manual",
     "Keep as the master list of what has been asked for. It is more complete than the "
     "open-questions register built from the other files."),
    ("Workbook A / Workbook B split",
     "A = the database. B = the analysis and risk model.",
     "Design Manual, and realised in practice",
     "Keep — and this is the same split as the warehouse-vs-analysis question in her latest "
     "note. She already decided it once."),
]


def main() -> None:
    audit = read_json(DATASETS / "workbook_audit.json")
    books = audit["workbooks"]
    max_tabs = max((w.get("sheet_count") or 0) for w in books)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- 1. the sheet she asked for: one file per row, one tab per column -----
    ws = wb.create_sheet("Workbook_Tab_Map")
    head = ["Workbook", "Purpose", "Family", "Build on?", "Authored", "Covers",
            "Years", "Tabs", "Rows"] + [f"Tab {i}" for i in range(1, max_tabs + 1)]
    for c, h in enumerate(head, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font = (HDR, HDRF) if c <= 9 else (TABF, Font(bold=True, size=9))
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"

    r = 1
    for w in books:
        r += 1
        yrs = w.get("years") or []
        ws.cell(r, 1, w["workbook"])
        ws.cell(r, 2, PURPOSE.get(w["workbook"], "")).alignment = Alignment(wrap_text=True,
                                                                           vertical="top")
        ws.cell(r, 3, w.get("family"))
        ws.cell(r, 4, "YES — build on this" if w.get("status", "").startswith("CURRENT")
                else ("same session — your call" if w.get("same_session_as_head") else ""))
        ws.cell(r, 5, w.get("authored") or "")
        ws.cell(r, 6, ", ".join(w.get("entities") or []))
        ws.cell(r, 7, f"FY{min(yrs) % 100:02d}–FY{max(yrs) % 100:02d}" if yrs else "")
        ws.cell(r, 8, w.get("sheet_count"))
        ws.cell(r, 9, w.get("rows_total"))
        for i, s in enumerate(w.get("sheets") or []):
            ws.cell(r, 10 + i, f"{s['name']} ({s['rows']})")
    for c, wdt in enumerate([46, 62, 22, 22, 12, 22, 14, 8, 8], 1):
        ws.column_dimensions[get_column_letter(c)].width = wdt
    for c in range(10, 10 + max_tabs):
        ws.column_dimensions[get_column_letter(c)].width = 30

    # ---- 2. the same thing on its side: one row per TAB -----------------------
    ws2 = wb.create_sheet("Tab_Index")
    ws2.cell(1, 1, "Sort by Tab to see which tabs recur across workbooks and which are "
                   "unique. A tab that appears in five files is a settled idea; one that "
                   "appears once is either a dead end or an unfinished good idea.")
    ws2.cell(1, 1).font = Font(italic=True, size=9, color="555555")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    for c, h in enumerate(["Tab", "Workbook", "Family", "Authored", "Rows", "Cols"], 1):
        cell = ws2.cell(2, c, h)
        cell.fill, cell.font = HDR, HDRF
    ws2.freeze_panes = "A3"
    rr = 2
    tab_counts: dict[str, int] = {}
    for w in books:
        for s in w.get("sheets") or []:
            rr += 1
            ws2.cell(rr, 1, s["name"])
            ws2.cell(rr, 2, w["workbook"])
            ws2.cell(rr, 3, w.get("family"))
            ws2.cell(rr, 4, w.get("authored") or "")
            ws2.cell(rr, 5, s["rows"])
            ws2.cell(rr, 6, s["cols"])
            tab_counts[s["name"]] = tab_counts.get(s["name"], 0) + 1
    for c, wdt in enumerate([38, 46, 22, 12, 8, 8], 1):
        ws2.column_dimensions[get_column_letter(c)].width = wdt

    # ---- 3. recurring tabs, ranked -------------------------------------------
    ws3 = wb.create_sheet("Recurring_Tabs")
    ws3.cell(1, 1, "How many workbooks each tab appears in. The top of this list is the "
                   "structure you kept coming back to — the strongest evidence of what the "
                   "final design should contain.")
    ws3.cell(1, 1).font = Font(italic=True, size=9, color="555555")
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    for c, h in enumerate(["Tab", "Appears in # workbooks", "Verdict"], 1):
        cell = ws3.cell(2, c, h)
        cell.fill, cell.font = HDR, HDRF
    rr = 2
    for name, n in sorted(tab_counts.items(), key=lambda x: (-x[1], x[0])):
        rr += 1
        ws3.cell(rr, 1, name)
        ws3.cell(rr, 2, n)
        ws3.cell(rr, 3, "Settled — carry forward" if n >= 4 else
                        ("Recurring" if n >= 2 else "Appears once — dead end or unfinished"))
    for c, wdt in enumerate([40, 22, 40], 1):
        ws3.column_dimensions[get_column_letter(c)].width = wdt

    # ---- 4. the decisions already made ---------------------------------------
    ws4 = wb.create_sheet("Decisions_Inventory")
    ws4.cell(1, 1, "Conventions and ideas already embedded in these workbooks — so the next "
                   "design carries them forward on purpose rather than by accident.")
    ws4.cell(1, 1).font = Font(italic=True, size=9, color="555555")
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    for c, h in enumerate(["Decision", "What it is", "Where it came from",
                           "Recommendation"], 1):
        cell = ws4.cell(2, c, h)
        cell.fill, cell.font = HDR, HDRF
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws4.freeze_panes = "A3"
    for i, (a, b, c_, d) in enumerate(DECISIONS, 3):
        for j, v in enumerate((a, b, c_, d), 1):
            cell = ws4.cell(i, j, v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c, wdt in enumerate([34, 46, 34, 66], 1):
        ws4.column_dimensions[get_column_letter(c)].width = wdt

    ws5 = wb.create_sheet("README")
    ws5.column_dimensions["A"].width = 24
    ws5.column_dimensions["B"].width = 110
    for i, (k, v) in enumerate([
        ("MFAS — Workbook Tab Map", ""),
        ("", ""),
        ("What this is", "Every Excel workbook in the project, one per row, with each tab in "
                         "its own column — as requested."),
        ("Workbook_Tab_Map", "The main sheet. Read left to right."),
        ("Tab_Index", "The same data one row per TAB, so you can sort by tab name."),
        ("Recurring_Tabs", "Which tabs recur across workbooks. The top of that list is the "
                           "structure you kept returning to."),
        ("Decisions_Inventory", "Conventions already embedded in these files, with a keep/"
                                "change recommendation for each."),
        ("⚠ GENERATED", "Rebuilt from the workbook audit on every run. Anything typed in here "
                        "is overwritten. Your own workbooks are never written to."),
        ("Row counts", "Counted by walking every row. Several of these files under-report "
                       "themselves as empty if you trust their stored dimensions."),
        ("Generated", date.today().isoformat()),
    ], 1):
        ws5.cell(i, 1, k).font = Font(bold=True, size=13 if i == 1 else 10)
        ws5.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
    wb.move_sheet("README", offset=-(len(wb.sheetnames) - 1))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"  wrote {OUT.relative_to(DATA.parent)}  "
          f"({OUT.stat().st_size / 1024:.0f} KB, {len(books)} workbooks, "
          f"{max_tabs} tab columns)")
    print(f"  {len(tab_counts)} distinct tab names across all workbooks")
    print("  most-recurring tabs:")
    for name, n in sorted(tab_counts.items(), key=lambda x: (-x[1], x[0]))[:8]:
        print(f"      {n} workbooks   {name}")


if __name__ == "__main__":
    main()
