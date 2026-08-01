"""The MFAS house style for Excel exports — one identity, shared by every workbook.

Why this is a module and not a handful of constants inside `s101`: there are two
exports (the warehouse and the tab map) and there will be more, and a look that is
re-typed per file stops being a look the first time someone edits one of them.

WHERE THE COLOURS COME FROM — this is the load-bearing part.

Nothing here was picked because it looked nice. Every colour is lifted from the
website's LIGHT theme in `assets/style.css`, where each one already carries a measured
contrast ratio against the page surface (there is a validator, `scripts/validate_palette.js`,
and the CSS records the failures that produced the current values). Excel is a light
surface, so the light set is the correct one — taking the DARK set, which is the site's
default identity, would land near-white text on white paper.

That shared origin is the whole point of the exercise. A resident who opens the site and
an analyst who opens the workbook should see the same hand. Copying the hex values into
this file rather than importing them is deliberate: the CSS is not machine-readable from
here, so the pairing is asserted by `tests/` instead, which fails if they drift.

WHAT MAKES IT LOOK UNLIKE EVERY OTHER SPREADSHEET

  * Gridlines OFF, and hairlines drawn where they actually mean something. Excel's grid
    is a drawing aid that got published by accident; a designed table rules the header
    and bands the body and leaves the rest quiet.
  * The header band is near-black with a mark-blue rule under it. Every stock Excel
    theme heads a table in a mid-blue or a grey. Ink plus one accent line is the site's
    grammar, and it is instantly not-Excel.
  * Warm paper (#FCFCFB), not #FFFFFF, on the sheets people read rather than filter.
  * A blue rule down the left edge of the narrative sheets — the wordmark, structurally.
  * Tab colours that mean something: blue dimensions, ink facts, orange coverage and
    quality, muted reference.

WHAT IS DELIBERATELY NOT DONE

  * No money format is ever applied by guessing at a column name. `Fact_Metric` and
    `Fact_Published_Figures` hold mixed units in one column — a tax rate of 0.6264 cents
    per $100 formatted as currency reads as "1" and that is a published falsehood, not a
    cosmetic slip. Callers name their money columns explicitly or get none.
  * No cell-by-cell fill on the large sheets. Banding is ONE conditional-format rule per
    sheet; 12,000 rows x 16 columns of explicit fills would triple the file for a stripe.
  * Nothing here writes a date, a random value or an iteration over a set. The exports
    must rebuild byte-identically (`common.normalise_xlsx`, and a test), so every
    ordering in this module is stable.
"""
from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- palette: assets/style.css, :root light theme --------------------------------
INK = "0B0B0B"        # --text-primary
INK_SOFT = "52514E"   # --text-secondary
MUTED = "6F6E6A"      # --text-muted
PAPER = "FCFCFB"      # --surface-1
PAPER_ALT = "F2F2EF"  # --surface-3  (the band)
RULE = "E1E0D9"       # --grid
RULE_FIRM = "C3C2B7"  # --axis

BLUE = "2A78D6"       # --accent / --series-1   (the mark)
BLUE_TEXT = "1F66BD"  # --accent-text           (5.25:1 — the one safe for text)
ORANGE = "D9541C"     # --series-2
GREEN = "0D8F66"      # --series-3

GOOD = "067806"
WARNING = "8A6400"
SERIOUS = "B3511F"
CRITICAL = "BD3232"

WHITE = "FFFFFF"

# Tints for the provenance chips. Mixed toward paper so a whole column of them stays
# readable — a saturated fill behind 12,000 rows of black text is unreadable and looks
# like a warning when it is only a category.
TINT_BLUE = "E4EEFA"
TINT_GREEN = "E2F1EB"
TINT_AMBER = "F6EFDD"
TINT_GREY = "EFEFEC"
TINT_RED = "F8E6E6"

MONEY = "#,##0;(#,##0)"      # accounting negatives — she reads statements, not JSON
MONEY_CENTS = "#,##0.00;(#,##0.00)"

HDR_FILL = PatternFill("solid", fgColor=INK)
HDR_FONT = Font(bold=True, color=WHITE, size=10, name="Calibri")
HDR_RULE = Border(bottom=Side(style="medium", color=BLUE))

BODY_FONT = Font(size=10, name="Calibri", color=INK)
NOTE_FONT = Font(italic=True, size=9, color=MUTED, name="Calibri")
KEY_FONT = Font(bold=True, size=10, color=INK, name="Calibri")
WORDMARK_FONT = Font(bold=True, size=26, color=INK, name="Calibri")
SUB_FONT = Font(size=11, color=INK_SOFT, name="Calibri")
LINK_FONT = Font(size=10, color=BLUE_TEXT, underline="single", name="Calibri")

PAPER_FILL = PatternFill("solid", fgColor=PAPER)
MARK_FILL = PatternFill("solid", fgColor=BLUE)

# Tab colours by family. The families are matched on the tab name, longest prefix
# first, so `Fact_Statement_Line` cannot be caught by a shorter rule.
_FAMILY = (
    ("Coverage_", ORANGE),
    ("Data_Quality", ORANGE),
    ("Open_Questions", ORANGE),
    ("Fact_", INK),
    ("Dim_", BLUE),
    ("Metric_Registry", BLUE),
    ("Permanent_IDs", BLUE),
    ("Source_Register", GREEN),
    ("Data_Dictionary", MUTED),
    ("Change_Log", MUTED),
    ("README", INK),
    ("Index", INK),
)

# Numbered analysis tabs ("1.0 Fact_...") are facts too; the digit prefix would
# otherwise fall through to no colour at all.
def tab_colour(name: str) -> str:
    stripped = name.lstrip("0123456789. ")
    for prefix, colour in _FAMILY:
        if stripped.startswith(prefix) or name.startswith(prefix):
            return colour
    return MUTED


def style_header(ws, row: int, ncols: int) -> None:
    """Ink band, white text, mark-blue rule beneath."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = HDR_RULE
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24


def band(ws, header_row: int, last_row: int, ncols: int) -> None:
    """Zebra the body with a single conditional-format rule.

    One rule for the whole range rather than a fill per cell: the warehouse tab alone
    is 12,000 rows, and explicit fills there cost more file than every figure in it.
    """
    if last_row <= header_row:
        return
    rng = f"A{header_row + 1}:{get_column_letter(ncols)}{last_row}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"MOD(ROW()-{header_row},2)=0"],
        fill=PatternFill("solid", bgColor=PAPER_ALT), stopIfTrue=False))


# The provenance vocabulary, given a colour so a column of it can be read at a glance.
# These are the project's own doctrine made visible: how a figure was obtained, and how
# far it is trusted. Order is fixed — Excel applies the first matching rule.
PROVENANCE_CHIPS = (
    ("digital-text", TINT_BLUE, BLUE_TEXT),
    ("ocr-arithmetic-verified", TINT_AMBER, WARNING),
    ("workbook-import", TINT_GREY, MUTED),
    ("derived", TINT_GREY, MUTED),
)
CONFIDENCE_CHIPS = (
    ("High", TINT_GREEN, GOOD),
    ("Medium", TINT_AMBER, WARNING),
    ("Working", TINT_GREY, MUTED),
    ("Pending", TINT_RED, CRITICAL),
)


def chips(ws, col_letter: str, header_row: int, last_row: int, mapping) -> None:
    """Colour-code an exact-match categorical column."""
    if last_row <= header_row:
        return
    rng = f"{col_letter}{header_row + 1}:{col_letter}{last_row}"
    for value, fill, text in mapping:
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'EXACT(${col_letter}{header_row + 1},"{value}")'],
            fill=PatternFill("solid", bgColor=fill),
            font=Font(color=text, bold=True, size=10, name="Calibri"),
            stopIfTrue=True))


def finish(ws, header_row: int, last_row: int, ncols: int, *, landscape=False) -> None:
    """Gridlines off, filter and freeze on, and a print setup that a person could use."""
    ws.sheet_view.showGridLines = False
    if last_row > header_row and ncols:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_row}"
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.oddFooter.left.text = "MFAS · Orange County Efficiency && Accountability Initiative"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = MUTED


def money(ws, header_row: int, last_row: int, col_idx: int, fmt: str = MONEY) -> None:
    """Accounting format on a column the CALLER has confirmed is single-unit.

    Never inferred from a header name. `Fact_Metric.Value` holds tax rates in cents per
    $100 beside dollar balances; formatting that column as currency would render 0.6264
    as "1" on a public artefact.
    """
    for r in range(header_row + 1, last_row + 1):
        ws.cell(r, col_idx).number_format = fmt


def narrative(ws, rows: int, *, bar_rows: int | None = None, cols: int = 7) -> None:
    """Warm paper and the left-edge mark bar, for sheets that are read not filtered.

    `cols` is the width of the paper, and it must match the sheet's actual last used
    column: paper drawn past the content reads as an empty panel, and paper stopping
    short of it leaves a white gutter down the right-hand side.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 0.9
    for r in range(1, (bar_rows or rows) + 1):
        ws.cell(r, 1).fill = MARK_FILL
    for r in range(1, rows + 1):
        for c in range(2, cols + 1):
            if not ws.cell(r, c).fill.fgColor.rgb or ws.cell(r, c).fill.patternType is None:
                ws.cell(r, c).fill = PAPER_FILL
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.left.text = "MFAS · Orange County Efficiency && Accountability Initiative"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED


def set_properties(wb, title: str, description: str, stamp: str) -> None:
    """Document properties — what Explorer, SharePoint and Drive show without opening it.

    `created`/`modified` are left alone on purpose: `common.normalise_xlsx` pins them to
    the ZIP epoch so two rebuilds are byte-identical, and writing a real time here would
    silently defeat it.
    """
    p = wb.properties
    p.title = title
    p.subject = "Municipal Financial Analysis System — Town of Hillsborough & Orange County, NC"
    p.creator = "Orange County Efficiency & Accountability Initiative"
    p.lastModifiedBy = "MFAS pipeline"
    p.category = "Municipal finance — published figures with document-level provenance"
    p.keywords = ("MFAS; Hillsborough; Orange County; North Carolina; municipal finance; "
                  "budget; ACFR; fiscal transparency")
    p.description = f"{description}\nBuild: {stamp}\nhttps://oc-accountability.github.io/MFAS/"
