"""Stage 101 — export everything to one Excel workbook, in Amy's own schema.

Her request:

    "One thing I will be looking for... is to have a data repository with all the data
     organized in a similar format to my excel spreadsheet
     Hillsborough_Municipal_Finance_Database_V1... Add additional tabs for any new content."

The file she names does not exist under that name; the lineage is
`Orange_County_Municipal_Finance_Database_v1.0` → `v1.1` → `..._Data_Warehouse_v1.2` →
`v2.0` → `..._Financial_Information_System_v2.2_Foundation`. This export follows the
**v2.2 Foundation** conventions rather than v1.1, because v2.2 is her own later evolution
of exactly this idea: numbered fact tabs, historical series rather than one year, a source
catalog, and a review dashboard. Matching the older shape would hand her back a structure
she had already outgrown.

**Her conventions, followed literally:**
  * every Fact tab ends with `Source_ID` and `Confidence`
  * `Fiscal_Year_ID` is the leading key, formatted `FY2027`
  * `Confidence` uses her vocabulary — High / Medium / Working / Pending
  * `Entity_ID` distinguishes the governments, and this adds `ORG_HILLSBOROUGH` beside her
    `ORG_OC`, because her workbook is county-centric and most of this data is the town's
  * Source_Register, Permanent_IDs, Dim_Fiscal_Year, Dim_Fund, Data_Quality_Gaps,
    Change_Log and Index all keep her column headings

**This file is GENERATED, and that is the one thing she must know about it.** It is rebuilt
from the datasets on every run, so anything typed into it is destroyed by the next build.
Her own workbooks stay hers and are never written to — stage 96 reads them and this stage
does not touch them. The README tab says so in the first cell anyone will read, because a
generated file that looks editable is a trap.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import DATA, DATASETS, read_json, build_stamp, normalise_xlsx  # noqa: E402
import workbook_style as st  # noqa: E402

OUT = DATA / "exports" / "MFAS_Data_Warehouse.xlsx"
TITLE_FONT = st.WORDMARK_FONT

# Columns that are unambiguously whole dollars, named one at a time.
# Deliberately NOT inferred from the header text: `Fact_Metric` and
# `Fact_Published_Figures` carry a Value column of mixed units — a tax rate of 0.6264
# cents per $100 rendered in an accounting format reads as "1", which is a published
# falsehood rather than a cosmetic slip. A tab absent from here simply gets no format.
MONEY_COLUMNS = {
    "Fact_Financial": ("Amount",),
    "Fact_Statement_Line": ("Amount",),
    "4.0 Fact_Capital_Projects": ("Amount",),
    "5.0 Fact_Requests_Declined": ("Amount",),
    "6.0 Fact_Requests_Funded": ("Amount",),
}


def ds(name):
    p = DATASETS / f"{name}.json"
    return read_json(p) if p.exists() else {}


def fy(v):
    """Her fiscal-year key format — TWO digits.

    Corrected 2026-07-28. This exported FY2027 because it was built to her Orange County
    v2.2 workbook. Her Hillsborough database — the one she actually named — uses FY27, and
    a key that does not match hers will not join to her data.
    """
    if isinstance(v, int):
        return f"FY{v % 100:02d}"
    return v or ""


def source_confidence(doc: dict) -> str:
    """Confidence in Amy's vocabulary, from WHO PUBLISHED IT and how it was read.

    The old rule was `"Pending" if scan else "High"` — it treated machine-readability
    as authority, so the initiative's own request workbook and a town audit both came
    out `High`. Readability is not authority. A perfectly parseable spreadsheet that
    this project wrote is `Working` until an official source confirms it; a scanned
    government audit is `Pending` because we decline to read it, not because the town
    is unreliable.
    """
    if (doc.get("source_authority") or "") != "government":
        return "Working"
    if doc.get("text_layer") == "scan":
        return "Pending"
    return "High"


def fact_confidence(fact: dict, authority_by_doc: dict) -> str:
    """Same principle for a published figure: authority first, then extraction."""
    auth = authority_by_doc.get(fact.get("source_doc"), "unknown")
    if auth != "government":
        return "Working"
    ext = fact.get("extraction")
    if ext == "derived":
        return "Medium"
    if ext == "ocr-arithmetic-verified":
        # Proven by the page's own arithmetic, which makes an undetected misread
        # very unlikely — but it is recognition, not a direct read.
        return "Medium"
    return "High"


def sheet(wb, title, headers, rows, note=None, widths=None):
    """One data tab in the house style.

    The header row stays at row 1 or 2 — never lower. `tests/` looks for the row
    carrying `Confidence` within the first three rows, and more to the point a reader
    scrolling a 12,000-row tab needs the frozen header to be the thing at the top, not
    a title block.
    """
    name = title[:31]
    ws = wb.create_sheet(name)
    hdr = 1
    if note:
        span = max(len(headers), 2)
        ws.cell(1, 1, note).font = st.NOTE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        ws.cell(1, 1).alignment = Alignment(vertical="top", wrap_text=True)
        hdr = 2
    for c, h in enumerate(headers, 1):
        ws.cell(hdr, c, h)
    st.style_header(ws, hdr, len(headers))

    r = hdr
    for row in rows:
        r += 1
        for c, v in enumerate(row, 1):
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            ws.cell(r, c, v)
    last = r

    for c, h in enumerate(headers, 1):
        w = widths[c - 1] if widths and c - 1 < len(widths) else min(
            42, max(12, len(str(h)) + 4,
                    *(len(str(row[c - 1])) + 2 for row in rows[:60]
                      if c - 1 < len(row) and row[c - 1] is not None) or [12]))
        ws.column_dimensions[get_column_letter(c)].width = w

    if note:
        # Excel does not auto-fit a MERGED cell, so a fixed height silently clips a long
        # note — and these notes carry the caveats, which is the worst thing to crop.
        # Measured from the REAL column widths, which is why this runs after the sizer:
        # the first attempt assumed every column was at the 42-unit cap, computed one
        # line for a note that wrapped to three, and clipped exactly as before.
        total_w = sum(ws.column_dimensions[get_column_letter(c)].width or 12
                      for c in range(1, len(headers) + 1))
        per_line = max(40, int(total_w * 1.15))   # ~1.15 chars per width unit at 9pt
        lines = min(6, max(1, -(-len(note) // per_line)))
        ws.row_dimensions[1].height = 12.5 * lines + 6

    st.band(ws, hdr, last, len(headers))
    # The provenance vocabulary, coloured wherever it appears. This is the one piece of
    # decoration that is also doctrine: how a figure was obtained and how far it is
    # trusted should be legible without reading the column.
    for label, mapping in (("Extraction", st.PROVENANCE_CHIPS),
                           ("Confidence", st.CONFIDENCE_CHIPS)):
        if label in headers:
            st.chips(ws, get_column_letter(headers.index(label) + 1), hdr, last, mapping)
    for col in MONEY_COLUMNS.get(name, ()):
        if col in headers:
            st.money(ws, hdr, last, headers.index(col) + 1)

    st.finish(ws, hdr, last, len(headers), landscape=len(headers) >= 8)
    ws.sheet_properties.tabColor = st.tab_colour(name)
    return ws


def cover(wb, index_rows, facts):
    """The first thing anyone sees — and the only sheet here that is purely identity.

    It earns its place by carrying the three things a reader needs before they trust a
    single cell: what this is, how big it is, and the KEY to the provenance colours used
    on every other tab. A cover that is only a logo is a wasted sheet.

    ⚠ THE COUNTS ON A COVER ARE STILL PUBLISHED FIGURES. The first draft printed
    "83 published figures" — `len(facts)`, which is only the Fact_Published_Figures tab —
    against a real total of 23,569, and "26 tabs" for a workbook with 28. A cover is
    exactly where a wrong number is least likely to be checked and most likely to be
    quoted, so both now come from the same places the rest of the file does: the
    coverage dataset, and the workbook's own sheet list.
    """
    cov = ds("coverage")
    ws = wb.create_sheet("MFAS")
    # Two columns only. An earlier version put the values in D with a spacer at C, and
    # the sheet then ran wider than a printable page — the entire right-hand column fell
    # off page 1 of the PDF and the cover rendered as a list of labels with no values.
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 84

    ws.cell(2, 2, "MFAS").font = st.WORDMARK_FONT
    ws.cell(3, 2, "Municipal Financial Analysis System").font = st.SUB_FONT
    ws.cell(4, 2, "Town of Hillsborough  ·  Orange County, North Carolina").font = st.SUB_FONT
    for c in (2, 3):
        ws.cell(6, c).border = Border(bottom=Side(style="medium", color=st.BLUE))
    ws.row_dimensions[2].height = 36

    rows = [
        ("What this is", "Every figure this project publishes, in one file — each one "
                         "traceable to a named document and, where the source prints one, "
                         "a page."),
        ("The rule", "A figure that cannot be traced to a source document is not published "
                     "at all. The build fails rather than guess."),
        ("Tabs", f"{len(wb.sheetnames)} — see Index. Colour-coded: blue dimensions, black "
                 f"facts, green sources, orange coverage and open questions."),
        ("Facts", (f"{cov.get('facts_total'):,} published figures, from "
                   f"{cov.get('documents_contributing')} of "
                   f"{cov.get('documents_total')} documents in the archive"
                   if cov.get("facts_total") else
                   f"{len(facts):,} in Fact_Published_Figures")),
        ("Still unread", (f"{cov.get('backlog_count')} documents — listed in "
                          f"Coverage_By_Document. A gap is published, not hidden."
                          if cov.get("backlog_count") else "")),
        ("Build", build_stamp()),
        ("", ""),
        ("KEY — how a figure was read", ""),
        ("digital-text", "Read straight from a document with real text. No character "
                         "recognition involved."),
        ("ocr-arithmetic-verified", "Recovered from a photograph of paper, and published "
                                    "only where its column adds up exactly to the total "
                                    "printed beside it."),
        ("workbook-import", "Imported from an analysis workbook and checked against an "
                            "official source."),
        ("", ""),
        ("KEY — how far it is trusted", ""),
        ("High", "An official published document."),
        ("Medium", "Derived here from official figures, or recovered by recognition."),
        ("Working", "From an analysis workbook, pending official confirmation."),
        ("Pending", "Awaiting a source. Not publishable on its own."),
        ("", ""),
        ("Website", "https://oc-accountability.github.io/MFAS/"),
        ("Everything", "https://github.com/oc-accountability/MFAS"),
    ]
    chip = {k: (fill, text)
            for k, fill, text in (*st.PROVENANCE_CHIPS, *st.CONFIDENCE_CHIPS)}
    r = 8
    for key, val in rows:
        if key.startswith("KEY —"):
            ws.cell(r, 2, key).font = Font(bold=True, size=10, color=st.BLUE_TEXT,
                                           name="Calibri")
        elif key in chip:
            fill, text = chip[key]
            c = ws.cell(r, 2, key)
            c.font = Font(bold=True, size=10, color=text, name="Calibri")
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif key:
            ws.cell(r, 2, key).font = st.KEY_FONT
        v = ws.cell(r, 3, val)
        v.alignment = Alignment(wrap_text=True, vertical="top")
        v.font = st.LINK_FONT if val.startswith("https://") else st.BODY_FONT
        if val.startswith("https://"):
            v.hyperlink = val
        if len(val) > 84:
            ws.row_dimensions[r].height = 30
        r += 1
    st.narrative(ws, r, cols=3)
    return ws


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    facts = ds("facts").get("facts", [])
    docs = ds("documents").get("documents", [])
    authority_by_doc = {d["id"]: d.get("source_authority") or "unknown" for d in docs}
    metrics = ds("metrics").get("metrics", {})
    wb = Workbook()
    wb.remove(wb.active)
    index_rows = []

    def add(title, headers, rows, purpose, note=None, widths=None):
        sheet(wb, title, headers, rows, note=note, widths=widths)
        index_rows.append([title[:31], purpose, len(rows)])

    # ---- README — the warning has to be the first thing read ------------------
    ws = wb.create_sheet("README")
    lines = [
        ("MFAS Data Warehouse", ""),
        ("", ""),
        ("⚠ GENERATED FILE",
         "Rebuilt from the published datasets on every run of `make etl`. ANYTHING YOU TYPE IN "
         "HERE WILL BE DESTROYED by the next build. Treat it as a export to read from and copy "
         "out of, not a workbook to author in."),
        ("Your workbooks are safe",
         "This project READS your workbooks and never writes to them — there is a test that "
         "fails the build if that ever changes. Your authored files stay yours."),
        ("Purpose",
         "Everything the project publishes, in one place, in the schema of your "
         "Hillsborough_Municipal_Financial_Database — Organization_ID, FY## keys, "
         "Source_ID + Source_Detail. CORRECTED 2026-07-28: the first version of this "
         "export followed your Orange County v2.2 workbook instead, which was the wrong "
         "one to copy for town data."),
        ("Schema",
         "Every Fact tab ends with Source_ID and Confidence. Fiscal_Year_ID is FY####. "
         "Organization_ID separates the two governments: ORG_HB and ORG_OC — your own IDs."),
        ("Confidence", "High = an official published document. Medium = derived by this project "
                       "from official figures. Working = imported from an analysis workbook. "
                       "Pending = awaiting a source."),
        ("Where the numbers come from",
         "Source_Register lists every document with its SHA-256. A figure that cannot be traced "
         "to one of them is not published at all."),
        ("Generated", build_stamp()),
        ("Live site", "https://oc-accountability.github.io/MFAS/"),
        ("Repository", "https://github.com/oc-accountability/MFAS"),
    ]
    # Column A is the mark bar, so the content shifts one right. `narrative()` fills it.
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 104
    for i, (k, v) in enumerate(lines, 1):
        cell = ws.cell(i, 2, k)
        cell.font = TITLE_FONT if i == 1 else st.KEY_FONT
        cell.alignment = Alignment(vertical="top")
        c = ws.cell(i, 3, v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = st.BODY_FONT
        if k.startswith("⚠"):
            cell.font = Font(bold=True, size=10, color=st.CRITICAL, name="Calibri")
        if v.startswith("https://"):
            c.font = st.LINK_FONT
            c.hyperlink = v
        ws.row_dimensions[i].height = 34 if len(v) > 120 else None
    ws.row_dimensions[1].height = 36
    st.narrative(ws, len(lines), cols=3)
    index_rows.append(["README", "What this file is, and the warning that it is generated", len(lines)])

    # ---- Change_Log ----------------------------------------------------------
    add("Change_Log", ["Version", "Date", "Change", "Notes"], [
        ["v1.0 export", build_stamp(),
         "First generated export of the MFAS pipeline into the v2.2 Foundation schema",
         "Built at Amy's request. Adds the tabs that did not exist in her workbook: tax-rate "
         "history, revenue by source, capital projects, funded/declined requests, utility "
         "block rates, the structural measures and the open-questions register."],
    ], "Version history of this export")

    # ---- Source_Register — her columns ---------------------------------------
    src_rows = []
    for d in sorted(docs, key=lambda x: (x.get("jurisdiction") or "", x["filename"])):
        scan = d.get("text_layer") == "scan"
        src_rows.append([
            d["id"],
            # Read, never inferred. The previous rule tested for the substring
            # "Orange" in the jurisdiction, which published ten initiative-authored
            # documents as Orange County government sources and thirteen Chapel Hill
            # documents as Town of Hillsborough.
            d.get("organization_id") or "ORG_UNKNOWN",
            d["filename"],
            fy(d.get("fiscal_year")) if d.get("fiscal_year") else "",
            (d.get("category") or d.get("format") or "").title(),
            f"{d.get('pages')} pages" if d.get("pages") else "",
            "Excluded from extraction — scanned" if scan else "Machine-readable source",
            source_confidence(d),
            (d.get("sha256") or "")[:16] + "…" if d.get("sha256") else "",
            d.get("source_authority") or "unknown",
        ])
    add("Source_Register",
        ["Source_ID", "Organization_ID", "Document", "Fiscal_Year", "Document_Type",
         "ACFR Page / Section", "Use in Model", "Confidence", "SHA-256 (first 16)",
         "Published_By"],
        src_rows, "Every source document, fingerprinted",
        note="A figure is only published if it traces to a row here. Scans are excluded because "
             "their hidden text scrambles digits.")

    # ---- Data_Dictionary -----------------------------------------------------
    add("Data_Dictionary", ["Table", "Field", "Definition", "Example", "Notes"], [
        ["All Fact Tables", "Organization_ID", "Permanent ID for the government", "ORG_HB",
         "Matches your Hillsborough database. ORG_OC is Orange County."],
        ["All Fact Tables", "Fiscal_Year_ID", "Fiscal year of the fact", "FY27",
         "June 30 year-end"],
        ["All Fact Tables", "Scenario", "Actual, Budget, Estimate, Projected, Recommended",
         "Budget", "Finer than Actual/Budget alone — the town publishes four bases"],
        ["All Fact Tables", "Source_ID", "Source register ID", "fy27-budget-message",
         "Traceable to Source_Register"],
        ["All Fact Tables", "Source_Page", "Page within the source document", "1",
         "Blank where the source states no page"],
        ["All Fact Tables", "Confidence", "High, Medium, Working, Pending", "High",
         "High = official document; Medium = derived here; Working = from an analysis workbook"],
        ["Fact_Published_Figures", "Metric", "Machine key for the measure",
         "county_property_tax_rate", "Defined in Metric_Registry"],
        ["Fact_Published_Figures", "Unit", "Unit of the value", "cents_per_100_valuation",
         "Never assume dollars — tax rates are cents per $100"],
    ], "Field conventions, matching your v2.2 Data_Dictionary")

    # ---- Permanent_IDs -------------------------------------------------------
    perm = [["ORG_HB", "Organization", "Town of Hillsborough", "", "Primary organization"],
            ["ORG_OC", "Organization", "Orange County", "", "County government (your ID)"]]
    for f, name, parent in [("FUND_GF", "General Fund", "ORG_HB"),
                            ("FUND_WS", "Water & Sewer Fund", "ORG_HB"),
                            ("FUND_SW", "Stormwater Fund", "ORG_HB")]:
        perm.append([f, "Fund", name, parent,
                     "Enterprise fund — paid by users, not taxes"
                     if f != "FUND_GF" else "Chief operating fund"])
    add("Permanent_IDs", ["ID", "ID_Type", "Name", "Parent_ID", "Definition / Notes"], perm,
        "Stable IDs, extending your scheme to the town")

    # ---- Dim_Fiscal_Year -----------------------------------------------------
    tco = ds("total_cost_of_ownership")
    years = sorted({f["fiscal_year"] for f in facts if f.get("fiscal_year")}
                   | {r["fiscal_year"] for r in tco.get("series", []) if r.get("fiscal_year")})
    reval = {2018, 2025, 2026}
    dim = []
    for y in years:
        audited = any(f["fiscal_year"] == y and f.get("basis") in ("actual", "audited")
                      for f in facts)
        dim.append([fy(y), fy(y), f"{y-1}-07-01", f"{y}-06-30",
                    "Closed" if audited else "Open",
                    "Audited Actual" if audited else "Budget / Projection",
                    "Complete" if audited else "In progress",
                    "Revaluation year / revenue-neutral rate applies" if y in reval else "",
                    "", ""])
    add("Dim_Fiscal_Year",
        ["Fiscal_Year_ID", "Fiscal Year", "Start Date", "End Date", "Financial Status",
         "Data Scenario", "Budget Status", "Revaluation Context", "Primary Actual Source",
         "Notes"], dim, "Fiscal years present in the data",
        note="Revaluation years matter: the rate falls because assessed values rise, so a "
             "falling rate is not a falling bill.")

    # ---- Metric registry -----------------------------------------------------
    add("Metric_Registry", ["Metric", "Label", "Unit", "Category", "Description"],
        [[k, v.get("label"), v.get("unit"), v.get("category"), v.get("description")]
         for k, v in sorted(metrics.items())],
        "What every metric key means")

    # ---- 1.0 the published figures ------------------------------------------
    add("1.0 Fact_Published_Figures",
        ["Organization_ID", "Fiscal_Year_ID", "Metric", "Value", "Unit", "Scenario",
         "Source_ID", "Source_Page", "Extraction", "Confidence"],
        [["ORG_OC" if f["metric"].startswith("county_") else "ORG_HB",
          fy(f.get("fiscal_year")), f["metric"], f.get("value"), f.get("unit"),
          (f.get("basis") or "").title(), f.get("source_doc"), f.get("source_page"),
          f.get("extraction"),
          fact_confidence(f, authority_by_doc)]
         for f in facts],
        "Every published figure, long format",
        note="This is the spine. Every number on the website is one of these rows.")

    # ---- 2.0 tax rate history ------------------------------------------------
    add("2.0 Fact_Tax_Rates_Hist",
        ["Fiscal_Year_ID", "County_Rate", "Town_Rate", "Combined_Rate",
         "School_District_Rate", "Fire_District_Rate", "Tax_On_Fixed_400k",
         "Corroborating_Editions", "Source_ID", "Confidence"],
        [[fy(r["fiscal_year"]), r.get("county_rate"), r.get("town_rate"),
          r.get("combined_rate"), r.get("school_district_rate"),
          r.get("fire_district_rate"), r.get("rate_on_fixed_400k"),
          r.get("corroborating_editions"),
          r.get("source", ""), "High" if r.get("corroborating_editions") else "Medium"]
         for r in tco.get("series", [])],
        "Both governments' tax rates, FY2013–FY2027",
        note="Dollars per $100 of assessed value. County rates come from ACFR Table 5, NOT "
             "Table 6 — Table 6's county column has a misplaced decimal in every edition held. "
             "Tax_On_Fixed_400k compares years on rate alone; it is NOT a bill history.")

    # ---- 3.0 revenue ---------------------------------------------------------
    rev = ds("revenue")
    rev_rows = []
    for y in rev.get("years", []):
        for k, v in (y.get("components") or {}).items():
            rev_rows.append([fy(y["fiscal_year"]), k, v,
                             (y.get("share_of_total") or {}).get(k),
                             y.get("basis"), y.get("state"),
                             "Hillsborough_GF_Trend_Schedules_v5", "Working"])
    add("3.0 Fact_GF_Revenue_Hist",
        ["Fiscal_Year_ID", "Revenue_Source", "Amount", "Share_Of_Total_Pct", "Scenario",
         "Reconciliation_State", "Source_ID", "Confidence"], rev_rows,
        "General Fund revenue by source",
        note="Shares are only filled where the components sum to the published total. Budget "
             "years reconcile exactly; audited years differ by up to $2.9M on a presentation "
             "difference that is a question for the town.")

    # ---- 4.0 capital projects ------------------------------------------------
    proj = ds("projects")
    add("4.0 Fact_Capital_Projects",
        ["Project_ID", "Project", "Organization_ID", "Fund", "Department", "Priority_Rank",
         "Plan_Window", "Total_Planned_Cost", "Creates_Recurring_Cost",
         "Recurring_Portion", "Funding_Unnamed", "Source_ID", "Source_Page", "Confidence"],
        [[p["project_id"], p["project_name"], "ORG_HB", p.get("fund"),
          p.get("department"), p.get("priority_rank"), p.get("plan_window"),
          p.get("total_planned_cost"), p.get("creates_recurring_cost"),
          (p.get("operating_budget_impact_quantified") or {}).get("recurring_portion"),
          any(f.get("unnamed_in_source") for f in p.get("funding_by_source", [])),
          p.get("source_doc"), (p.get("source_pages") or [None])[0], "High"]
         for p in proj.get("projects", [])],
        "The 27-project capital register",
        note="Every project reconciled to its own printed totals, per year column. "
             "Funding_Unnamed flags projects whose funding source the document does not name.")

    # ---- 5.0 / 6.0 requests --------------------------------------------------
    tr = ds("tradeoffs")
    for n, key, label in (("5.0", "declined", "declined"), ("6.0", "funded", "funded")):
        add(f"{n} Fact_Requests_{label.title()}",
            ["Fiscal_Year_ID", "Request", "Fund", "Department", "FY2027", "Three_Year_Total",
             "Impact_If_Not_Funded", "Source_ID", "Source_Page", "Confidence"],
            [["FY2027", r["request"], r.get("fund"), r.get("department"), r.get("fy2027"),
              r.get("total_three_year"), r.get("impact_if_not_funded"),
              "fy27-budget-and-financial-plan-recommended", r.get("source_page"), "High"]
             for r in tr.get(key, [])],
            f"New spending requests the town {label}",
            note=("The town publishes both lists. Impact_If_Not_Funded is the town's own words."
                  if key == "declined" else "The funded side of the same decision."))

    # ---- 7.0 utility block rates --------------------------------------------
    ur = ds("utility_rates")
    urows = []
    for name, rs in (ur.get("rate_sets") or {}).items():
        for basis in ("current", "recommended"):
            b = rs[basis]
            urows.append([fy(b["fiscal_year"]), rs["service"].title(), rs["location"].title(),
                          b["threshold_gallons"], b["block1_charge"], b["block2_per_1000"],
                          basis.title(), "fy27-budget-and-financial-plan-recommended", "High"])
    add("7.0 Fact_Utility_Block_Rates",
        ["Fiscal_Year_ID", "Service", "Location", "Threshold_Gallons", "Block1_Charge",
         "Block2_Per_1000", "Scenario", "Source_ID", "Confidence"], urows,
        "Water and sewer block rates",
        note="Block 1 covers the first N gallons; Block 2 is per 1,000 above it. This structure "
             "reproduces all eight increase figures the town states in prose.")

    # ---- 8.0 the structural measures ----------------------------------------
    structure = ds("structure")
    b = structure.get("reading_burden", {})
    sep = structure.get("run_separately", {})
    add("8.0 Fact_Structure_Measures",
        ["Measure", "Value", "Unit", "Notes", "Source_ID", "Confidence"], [
            ["Governments a resident must read", b.get("governments_a_resident_must_read"),
             "count", "To answer why a bill went up", "manifest", "High"],
            ["Documents, current budget cycle", b.get("current_cycle_documents"), "count", "",
             "manifest", "High"],
            ["Pages, current budget cycle", b.get("current_cycle_pages"), "pages",
             "A floor — only documents this pipeline could measure", "manifest", "High"],
            ["Town administration (broad)", (sep.get("administration_broad") or {}).get("total"),
             "USD", "Share of General Fund: "
             f"{(sep.get('administration_broad') or {}).get('share_of_general_fund_pct')}%",
             "lineitems", "Medium"],
            ["Town administration (narrow)", (sep.get("administration_narrow") or {}).get("total"),
             "USD", "Excludes Communications and Facility Management — the boundary is arguable",
             "lineitems", "Medium"],
            ["County administration", None, "USD",
             "NOT EXTRACTED — publishing the town's figure alone would invite a comparison "
             "against nothing", "", "Pending"],
        ], "What it costs a resident to answer one question",
        note="Poses the structural question. It does NOT answer whether two administrations cost "
             "more than one — no document in the archive compares them.")

    # ---- Amy's own imported analysis ----------------------------------------
    wbb = ds("workbook_b")
    add("9.0 Fact_Change_Drivers",
        ["Driver", "Amount", "Cents_Equivalent", "Period", "Budget_Category", "Commentary",
         "Follow_Up_Question", "Source_ID", "Confidence"],
        [[m["driver"], m.get("amount"), m.get("cents_equivalent"), m.get("period"),
          m.get("budget_category"), m.get("commentary"), m.get("follow_up_question"),
          m.get("source"), (m.get("confidence") or "Working")]
         for m in wbb.get("material_change_drivers", [])],
        "Your Material Change Drivers, imported",
        note="Imported from your v5 Audit Edition and never modified. Cents_Equivalent is "
             "recomputed here from the town's published $240,000 per cent.")

    # ---- Data_Quality_Gaps + the register, in her columns -------------------
    reg = ds("questions")
    gaps, qrows = [], []
    for r in reg.get("register", []):
        qrows.append([r["id"], r["owner"], r["topic"], r["question"], r["status"],
                      r.get("why_it_matters"), r.get("answer"), r.get("raised_by")])
        if r["status"] != "answered":
            gaps.append([r["id"],
                         "High" if r["owner"] in ("town", "county") else "Medium",
                         r["topic"], r["status"].title(), r.get("why_it_matters") or "",
                         {"town": "Records request to the Town",
                          "county": "Records request to the County",
                          "amy": "Your decision",
                          "david": "Repository owner action",
                          "pipeline": "Work this project owes"}.get(r["owner"], "")])
    # ---- the warehouse core (stage 87) — her 2026-07-29 go-ahead --------------
    # The pipeline is the system of record and these tabs are the generated view:
    # her decision, quoted in the register. The government is a COLUMN here — no
    # tab or file carries a municipality's name, which is the whole point.
    wh = ds("warehouse")
    if wh:
        add("Dim_Organization",
            ["Organization_ID", "Name", "Status"],
            [[d["Organization_ID"], d["Name"], d["Status"]]
             for d in wh["dim_organization"]],
            "Frozen. Future IDs reserved before their data arrives — her pattern.",
            note="one row per organization; Status=Future means the ID is reserved, "
                 "her Hillsborough_Municipal_Financial_Database_v1 convention")
        add("Dim_Scenario",
            ["Scenario", "Meaning"],
            [[d["Scenario"], d["meaning"]] for d in wh["dim_scenario"]],
            "Frozen. Five values; an Actual never overwrites an Estimate.",
            note="one row per scenario; appending a year's Actual keeps its Estimate — "
                 "what was projected vs what happened is the history nobody else keeps")
        add("Dim_Fiscal_Year_WH",
            ["Fiscal_Year_ID", "Scenarios loaded"],
            [[d["Fiscal_Year_ID"], ", ".join(d["scenarios_loaded"])]
             for d in wh["dim_fiscal_year"]],
            "Frozen. One row per year present in Fact_Financial.",
            note="one row per fiscal year loaded; a new budget year is one new row "
                 "here and appended fact rows — no new columns, no new tabs")
        add("Fact_Financial",
            wh["columns"],
            wh["rows"],
            "THE warehouse core: both governments, one table, one schema.",
            note="GRAIN: " + wh["grain"])
        fm = wh.get("fact_metric") or {}
        if fm.get("rows"):
            add("Fact_Metric",
                fm["columns"],
                [[r.get(c) for c in fm["columns"]] for r in fm["rows"]],
                "Facts that are not fund dollars — fund balance, net position, debt, "
                "capital, schools, outlook.",
                note="GRAIN: " + fm["grain"] + "  " + fm["why_separate"])
        fsl = wh.get("fact_statement_line") or {}
        if fsl.get("rows"):
            add("Fact_Statement_Line",
                fsl["columns"],
                [[r.get(c) for c in fsl["columns"]] for r in fsl["rows"]],
                "Verified and cited statement lines whose COLUMN MEANING is not yet "
                "established — read this tab knowing the basis is unknown.",
                note="GRAIN: " + fsl["grain"] + "  " + fsl["why_separate"])

    # Coverage travels WITH the workbook. A reader who never opens the JSON should
    # still be able to see which documents fed this and which did not — a coverage
    # claim that lives only in a commit message is a coverage claim nobody checks.
    cov = ds("coverage")
    if cov:
        add("Coverage_By_Document",
            ["Document", "Filename", "Jurisdiction", "Format", "Fiscal_Year",
             "Facts_Total", "Fact_Financial", "Fact_Metric", "Fact_Statement_Line",
             "Status", "Why"],
            [[r["document"], r["filename"], r["jurisdiction"], r["format"],
              r["fiscal_year"], r["facts_total"], r["Fact_Financial"],
              r["Fact_Metric"], r["Fact_Statement_Line"], r["status"], r["why"]]
             for r in cov["documents"]],
            f"{cov['documents_contributing']} of {cov['documents_total']} documents feed "
            f"{cov['facts_total']:,} facts. Only status='not-yet-read' is a gap in the work.",
            note=cov["how_to_read_this"])
        yrs = [k for k in cov["facts_by_org_and_year"][0] if k != "Organization_ID"]
        add("Coverage_By_Year",
            ["Organization_ID"] + yrs,
            [[m["Organization_ID"]] + [m[y] for y in yrs]
             for m in cov["facts_by_org_and_year"]],
            "Facts per government per year. A thin year is a real hole.",
            note="A thin year is a real hole. This note used to say Hillsborough FY2018-"
                 "FY2019 were thin for want of a digital audit and that the fix was not more "
                 "code — wrong twice: the town had sent everything it holds, and the gap was "
                 "two defects in this pipeline (a scan-selection rule that preferred the "
                 "smallest file, and line-banding that could not survive a 0.4° page "
                 "rotation). Fixing them took FY2019 from 81 verified statement lines to 464.")

    add("Data_Quality_Gaps",
        ["Gap_ID", "Priority", "Topic", "Current Status", "Why It Matters",
         "Likely Resolution / Next Steps"], gaps, "Open gaps, in your schema")
    add("Open_Questions",
        ["ID", "Owner", "Topic", "Question", "Status", "Why It Matters", "Answer", "Raised By"],
        qrows, "The full register, including answered items",
        note="Answered items are kept — a question that quietly vanishes is indistinguishable "
             "from one that was forgotten.")

    # ---- Index (her convention: last, listing every tab) ---------------------
    sheet(wb, "Index", ["Tab", "Purpose", "Rows"], index_rows,
          note="Every tab in this workbook. Tab colours group them: blue = dimensions, "
               "black = facts, green = sources, orange = coverage and open questions.",
          widths=[32, 78, 9])
    wb.move_sheet("Index", offset=-(len(wb.sheetnames) - 1))
    wb.move_sheet("README", offset=-(len(wb.sheetnames) - 1))
    cover(wb, index_rows, facts)
    wb.move_sheet("MFAS", offset=-(len(wb.sheetnames) - 1))

    st.set_properties(
        wb, "MFAS Data Warehouse — Hillsborough & Orange County, NC",
        "Every figure this project publishes, each traceable to a document and page.",
        build_stamp())
    wb.save(OUT)
    # Fixed ZIP timestamps: without this two rebuilds of identical data still
    # differ byte-for-byte, and `git diff --stat data/` stays permanently dirty.
    normalise_xlsx(OUT)
    size = OUT.stat().st_size
    print(f"  wrote {OUT.relative_to(DATA.parent)}  ({size / 1024:.0f} KB, "
          f"{len(wb.sheetnames)} tabs)")
    for t, purpose, n in index_rows:
        print(f"      {t:34} {n:6} rows   {purpose[:44]}")


if __name__ == "__main__":
    main()
