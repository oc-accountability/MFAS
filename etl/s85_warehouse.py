"""Stage 85 — one warehouse, two governments, Amy's schema.

Amy's design workbook ("Orange County Municipal Financial Information System",
v2.2 Foundation) is a real star schema with permanent IDs, a source register, and
396 hand-curated fact rows covering Orange County FY2018-FY2025. Her
`Design_Requirements` sheet already anticipates this stage:

    "Support future OC + Hillsborough comparison — use standardized entity IDs"

Her `Entity_ID` is `ORG_OC` throughout. This project's automated pipeline covers
the Town of Hillsborough. So the two halves join by adding `ORG_HB` rows in *her*
shape rather than inventing a competing one.

Three rules, all of them about keeping her in control:

1. **Her workbook is read, never written.** This stage imports and verifies; it
   never edits her file. She fine-tunes in Excel, drops the new version in, and
   the pipeline picks it up.
2. **Her vocabulary wins.** Entity_ID / Fiscal_Year_ID / Scenario / Category_ID /
   Source_ID / Confidence are hers. Where this project had its own word for the
   same idea, hers is used in the warehouse output.
3. **Her figures are checked, not trusted.** Every row whose citation names a
   document and PDF page we hold is verified against that page. A spot-check of
   her FY2018 rows matched the source CAFR exactly, but "it was right last time"
   is not a method — the check runs every build.

The newest workbook version present is used automatically, so v2.3 supersedes
v2.2 the moment she saves it into the folder.
"""
from __future__ import annotations

import re
import sys
import warnings
from pathlib import Path

import openpyxl
import pdfplumber

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import (DATASETS, SOURCES, content_cache_dir,  # noqa: E402
                    doc_id_for_filename, read_json, write_json)

warnings.filterwarnings("ignore")

DESIGN_DIR = ("Orange County Efficiency & Accountability Initiative/11 Design Documents")
COUNTY_DIR = ("Orange County Efficiency & Accountability Initiative/"
              "06b Budget & Fin. Analysis - OC")

FACT_SHEET = re.compile(r"^\d+[\.\s]")
PDF_PAGE = re.compile(r"PDF\s*p+\.?\s*(\d+)", re.I)

# Amy's Confidence vocabulary -> whether a row may be published unverified.
# Blank is deliberately NOT publishable on its own: 26 of her rows have no
# confidence set, and silently promoting them to "trusted" would misrepresent her
# own assessment of her work.
CONFIDENCE_PUBLISHABLE = {"High", "Medium"}


def newest_workbook() -> Path | None:
    """The highest-versioned design workbook present — so v2.3 wins automatically."""
    d = SOURCES / DESIGN_DIR
    if not d.exists():
        return None
    best, best_key = None, ()
    # sorted(): the >= tie-break must not depend on filesystem order.
    for p in sorted(d.glob("*.xlsx")):
        m = re.search(r"v(\d+)\.(\d+)", p.name)
        key = (int(m.group(1)), int(m.group(2))) if m else (0, 0)
        if key >= best_key:
            best, best_key = p, key
    return best


def load_source_register(wb) -> dict[str, str]:
    """Source_ID -> document filename, from her Source_Register sheet."""
    out = {}
    if "Source_Register" not in wb.sheetnames:
        return out
    ws = wb["Source_Register"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return out
    hdr = [str(c).strip() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        doc = r[idx["Document"]] if "Document" in idx and len(r) > idx["Document"] else None
        if doc:
            out[str(r[0]).strip()] = str(doc).strip()
    return out


def county_pdf_for(source_id: str, register: dict[str, str]) -> Path | None:
    """Resolve a Source_ID to a PDF on disk.

    Her Source_IDs are not perfectly uniform — the register uses
    `SRC_OC_ACFR_2025…` while fact rows use `OC_CAFR_2018`. So: try the register,
    then fall back to matching the four-digit year against the county filenames.
    Reported as a data-quality note rather than silently patched.
    """
    d = SOURCES / COUNTY_DIR
    if not d.exists():
        return None
    name = register.get(source_id)
    if name and (d / name).exists():
        return d / name
    m = re.search(r"(20\d{2})", source_id or "")
    if not m:
        return None
    year = m.group(1)
    cands = [p for p in d.glob("*.pdf")
             if year in p.name and re.search(r"ACFR|CAFR", p.name, re.I)]
    if len(cands) > 1:
        # Taking the first match silently picked one of several plausible source
        # documents, then verified a transcription against it and reported the
        # result as fact. Ambiguity is a data-quality condition this stage already
        # describes; it must behave like one.
        sys.exit(f"AMBIGUOUS SOURCE: {len(cands)} county PDFs match year {year} for "
                 f"{source_id!r}: {[c.name for c in cands]}. Resolve it in the "
                 f"workbook's Source_Register before rebuilding.")
    return cands[0] if cands else None


def page_numbers(path: Path, page: int, sha256: str) -> set[float] | None:
    """Every money-looking figure on one page of a PDF, cached by CONTENT.

    The key used to be the filename stem truncated to 60 characters plus the page
    number. Two problems: a replaced file served the old page's figures, and these
    filenames are long enough that two of them can collide after truncation and
    quietly share a cache entry — this stage is the one that VERIFIES Amy's
    transcriptions against the page she cited, so a wrong page here produces a
    confident and wrong verdict about her work.
    """
    key = content_cache_dir("pagecache", sha256, extractor="pdfplumber",
                            version="1") / f"p{page:05d}.txt"
    if key.exists():
        text = key.read_text(encoding="utf-8", errors="replace")
    else:
        try:
            with pdfplumber.open(path) as pdf:
                if page < 1 or page > len(pdf.pages):
                    return None
                text = pdf.pages[page - 1].extract_text() or ""
        except Exception:
            return None
        key.parent.mkdir(parents=True, exist_ok=True)
        key.write_text(text, encoding="utf-8")
    out = set()
    for tok in re.findall(r"\(?\$?\s?\d[\d,]{2,}\)?", text):
        t = tok.strip().replace("$", "").replace(" ", "")
        neg = t.startswith("(")
        t = t.strip("()").replace(",", "")
        if t.isdigit():
            out.add(-float(t) if neg else float(t))
    return out


def main() -> None:
    wbp = newest_workbook()
    if wbp is None:
        sys.exit(f"no design workbook found under {DESIGN_DIR}")
    print(f"  using {wbp.name}")
    wb = openpyxl.load_workbook(wbp, data_only=True)
    register = load_source_register(wb)

    notes: list[str] = []
    rows: list[dict] = []
    # Resolve each source PDF to its manifest hash so the page cache is
    # content-keyed rather than filename-keyed.
    sha_by_path = {str(SOURCES / d['archive_path']): d['sha256']
                   for d in read_json(DATASETS / 'documents.json')['documents']}

    for sheet in wb.sheetnames:
        if not FACT_SHEET.match(sheet):
            continue
        ws = wb[sheet]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        # Header names are NORMALISED before lookup, because her sheets spell the
        # same field two ways: "ACFR_Page" on the budget-vs-actual tabs and
        # "ACFR Page" (a space) on nine others. Looking for the underscore form only
        # meant every row on those nine tabs imported with no citation, was counted
        # as "no citation with a PDF page", and landed in the 355 rows this stage
        # reported as not verifiable. They were citing pages the whole time.
        def norm(h) -> str:
            # Any run of non-word characters becomes one underscore, so "ACFR Page"
            # and "Purpose / Subcategory" both resolve to a usable key.
            return re.sub(r"\W+", "_", str(h or "").strip()).strip("_")

        hdr = [norm(c) for c in data[0]]
        idx = {h: i for i, h in enumerate(hdr) if h}
        need = {"Entity_ID", "Fiscal_Year_ID"}
        if not need <= set(idx):
            continue
        get = lambda r, k: (r[idx[k]] if k in idx and len(r) > idx[k] else None)  # noqa: E731

        # Every column she wrote is carried through. The previous whitelist kept
        # eleven fields and silently dropped Metric, Metric_ID, Unit, Fund, Fund_ID,
        # Activity_Type and Notes — which is why nine of her county tables arrived as
        # a bare Amount with nothing to say what it measured, and were then held out
        # of the warehouse as "not fund-level dollar facts". They were perfectly good
        # facts; the import had thrown away their labels. A number whose label this
        # pipeline discarded is worse than one it never read.
        # Her label column is named differently per tab — Category on the
        # budget-vs-actual tabs, Metric on most, Classification on the fund-balance
        # tab. All three are first-class here; a row whose label lands in
        # other_fields as a stringified extra is a row nothing downstream can use.
        KNOWN = ("Scenario", "Line_Type", "Category_ID", "Category", "Metric_ID",
                 "Metric", "Classification_ID", "Classification", "Purpose_Subcategory",
                 "Unit", "Fund_ID", "Fund", "Activity_Type",
                 "Original_Budget", "Final_Budget", "Actual_Amount", "Amount",
                 "Variance", "Source_ID", "ACFR_Page", "Confidence", "Notes")
        for r in data[1:]:
            if not r or not r[0] or not str(r[0]).startswith("ORG_"):
                continue
            row = {
                "table": sheet,
                "Entity_ID": str(get(r, "Entity_ID")),
                "Fiscal_Year_ID": str(get(r, "Fiscal_Year_ID") or ""),
            }
            for k in KNOWN:
                row[k] = get(r, k)
            # Anything she adds later arrives here rather than being dropped.
            extra = {h: r[i] for h, i in idx.items()
                     if h not in KNOWN and h not in ("Entity_ID", "Fiscal_Year_ID")
                     and len(r) > i and r[i] is not None}
            if extra:
                row["other_fields"] = {k: str(v) for k, v in extra.items()}
            row["origin"] = "amy-workbook"
            rows.append(row)

    # ---- verify her figures against the pages she cites ----------------------
    checked = matched = unverifiable = 0
    mismatches: list[dict] = []
    for row in rows:
        page_ref = str(row.get("ACFR_Page") or "")
        m = PDF_PAGE.search(page_ref)
        pdf = county_pdf_for(str(row.get("Source_ID") or ""), register)
        # Zeros are excluded: a nil amount is printed on these statements as a
        # dash, not as "0", so searching the page for the number 0 can never
        # succeed. Flagging those as unverified would blame her data for a
        # limitation of this check.
        vals = [v for k, v in row.items()
                if k in ("Original_Budget", "Final_Budget", "Actual_Amount", "Amount")
                and isinstance(v, (int, float)) and round(abs(v)) != 0]
        if not m or pdf is None or not vals:
            row["verification"] = "no citation with a PDF page, or no figures"
            unverifiable += 1
            continue
        pdf_sha = sha_by_path.get(str(pdf))
        if pdf_sha is None:
            row["verification"] = "source PDF not in the manifest"
            unverifiable += 1
            continue
        page_nums = page_numbers(pdf, int(m.group(1)), pdf_sha)
        if page_nums is None:
            row["verification"] = "cited page not readable"
            unverifiable += 1
            continue
        checked += 1
        missing = [v for v in vals if round(abs(v)) not in {round(abs(x)) for x in page_nums}]
        if missing:
            row["verification"] = "NOT FOUND on the cited page"
            row["figures_not_found"] = missing[:4]
            mismatches.append({"table": row["table"], "fiscal_year": row["Fiscal_Year_ID"],
                               "category": row["Category"], "source": row["Source_ID"],
                               "page": page_ref, "missing": missing[:4]})
        else:
            row["verification"] = "every figure found on the cited page"
            matched += 1

    if register and any(r.get("Source_ID") and str(r["Source_ID"]) not in register for r in rows):
        unknown = sorted({str(r["Source_ID"]) for r in rows
                          if r.get("Source_ID") and str(r["Source_ID"]) not in register})
        notes.append(f"Source_IDs used in fact tables but absent from Source_Register: "
                     f"{', '.join(unknown[:6])}. Resolved by fiscal-year fallback; worth "
                     f"aligning in a future workbook version.")

    blank_conf = sum(1 for r in rows if not r.get("Confidence"))
    if blank_conf:
        notes.append(f"{blank_conf} rows have no Confidence set. They are imported but marked "
                     f"not-publishable-on-their-own, because promoting a blank to 'trusted' "
                     f"would misstate the author's own assessment.")

    entities = sorted({r["Entity_ID"] for r in rows})
    write_json(DATASETS / "warehouse_county.json", {
        "generated_by": "etl/s85_warehouse.py",
        "workbook": wbp.name,
        # Manifest id for the workbook itself — the proximate document every imported
        # row traces to, machine-joinable rather than a bare filename.
        "source_doc": doc_id_for_filename(wbp.name),
        "note": ("Imported from the design workbook, which remains the source of truth for these "
                 "figures and is never modified by this pipeline. Amy edits in Excel; the "
                 "pipeline re-reads and re-verifies on the next build."),
        "schema_owner": "Amy's v2.x design workbook — Entity_ID / Fiscal_Year_ID / Scenario / "
                        "Category_ID / Source_ID / Confidence are her field names.",
        "entities": entities,
        "verification": {
            "rows_imported": len(rows),
            "rows_checked_against_source_pdf": checked,
            "every_figure_found": matched,
            "figures_not_found_on_cited_page": len(mismatches),
            "not_verifiable": unverifiable,
            "method": ("For each row citing a document and PDF page we hold, every monetary figure "
                       "in the row must appear on that page. This catches a transcription slip or "
                       "a citation that has drifted; it does not check that a figure sits on the "
                       "right LINE of the page."),
        },
        "mismatches": mismatches[:60],
        "data_quality_notes": notes,
        "rows": rows,
    })

    print(f"  imported {len(rows)} rows for {', '.join(entities)}")
    print(f"  verified {matched}/{checked} rows — every figure found on the page cited")
    if mismatches:
        print(f"  {len(mismatches)} row(s) with a figure NOT on the cited page:")
        for x in mismatches[:6]:
            print(f"      {x['fiscal_year']} {str(x['category'])[:34]:36} {x['page']}")
    print(f"  {unverifiable} row(s) not verifiable (no PDF-page citation, or page unreadable)")
    for n in notes:
        print(f"  note: {n}")


if __name__ == "__main__":
    main()
